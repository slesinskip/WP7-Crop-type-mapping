import argparse
from pathlib import Path
import numpy as np
from osgeo import gdal
import geopandas as gpd
from sklearn.metrics import confusion_matrix, precision_recall_fscore_support
import openpyxl
from openpyxl.styles import Font

# Minimal mapping of track → country; extend as needed
TRACK_REGIONS = {
    'P1':  'AU', 'P1a': 'AU',
    'P2':  'IR',
    'P3':  'NL',
    'P4':  'PT', 'P4a': 'PT',
    'P5':  'XX', 'P5a': 'XX', 'P5b': 'XX',
    # add more as you create new tracks…
}

def find_masked_files(base_dir: Path, tr: str, country: str):
    """
    Look for [{country}_{tr}_classified_masked.tif,
              {country}_{tr}_confidence_masked.tif]
    in either:
      - classification_results/classification/
      - classification_results/
    Returns (cls_fp, conf_fp) or (None, None).
    """
    candidates = [
        base_dir / tr / 'classification_results' / 'classification',
        base_dir / tr / 'classification_results'
    ]
    cls_name  = f"{country}_{tr}_classified_masked.tif"
    conf_name = f"{country}_{tr}_confidence_masked.tif"
    for folder in candidates:
        cls_fp  = folder / cls_name
        conf_fp = folder / conf_name
        if cls_fp.exists() and conf_fp.exists():
            return cls_fp, conf_fp
    return None, None

def discover_tracks(base_dir: Path, prefix: str):
    """
    Automatically discover subfolders under base_dir whose
    names start with `prefix` and exist in TRACK_REGIONS.
    Returns list of (tr, country, cls_fp, conf_fp).
    """
    tracks = []
    for sub in base_dir.iterdir():
        tr = sub.name
        if not tr.startswith(prefix):
            continue
        country = TRACK_REGIONS.get(tr)
        if country is None:
            continue
        cls_fp, conf_fp = find_masked_files(base_dir, tr, country)
        if cls_fp:
            tracks.append((tr, country, cls_fp, conf_fp))
    return tracks

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--track', required=True,
                        help='Base track prefix (e.g. P1)')
    args = parser.parse_args()
    prefix = args.track

    base_dir = Path("D:/AIML_CropMapper/workingDir")
    tracks   = discover_tracks(base_dir, prefix)
    if not tracks:
        raise FileNotFoundError(
            f"No valid classification/confidence files for tracks starting with {prefix}"
        )

    print(f"Discovered tracks: {[t for t,_,_,_ in tracks]}")

    # --- compute union extent & grid ---------------------------------------
    ds0   = gdal.Open(str(tracks[0][2]))
    proj  = ds0.GetProjection()
    gt0   = ds0.GetGeoTransform()
    resX, resY = gt0[1], abs(gt0[5])

    extents = []
    for _, _, cls_fp, _ in tracks:
        ds = gdal.Open(str(cls_fp))
        gt = ds.GetGeoTransform()
        c, r = ds.RasterXSize, ds.RasterYSize
        minX = gt[0]
        maxY = gt[3]
        maxX = gt[0] + c * gt[1]
        minY = gt[3] + r * gt[5]
        extents.append((minX, maxY, maxX, minY))

    minX = min(e[0] for e in extents)
    maxY = max(e[1] for e in extents)
    maxX = max(e[2] for e in extents)
    minY = min(e[3] for e in extents)

    cols = int(np.ceil((maxX - minX) / resX))
    rows = int(np.ceil((maxY - minY) / resY))
    gt_global = (minX, resX, 0, maxY, 0, -resY)

    print(f"Global mosaic: {cols} cols × {rows} rows")

    # --- warp and stack -----------------------------------------------------
    class_stack = []
    conf_stack  = []
    for tr, country, cls_fp, conf_fp in tracks:
        # classification
        mem_cls = gdal.Warp(
            '', str(cls_fp), format='MEM',
            width=cols, height=rows,
            outputBounds=(minX, minY, maxX, maxY),
            dstSRS=proj,
            resampleAlg=gdal.GRA_NearestNeighbour
        )
        arr_cls = mem_cls.GetRasterBand(1).ReadAsArray()
        nod     = mem_cls.GetRasterBand(1).GetNoDataValue()

        # confidence
        mem_conf = gdal.Warp(
            '', str(conf_fp), format='MEM',
            width=cols, height=rows,
            outputBounds=(minX, minY, maxX, maxY),
            dstSRS=proj,
            resampleAlg=gdal.GRA_NearestNeighbour
        )
        arr_conf = mem_conf.GetRasterBand(1).ReadAsArray().astype(np.float32)

        # mask out nodata
        if nod is not None:
            arr_conf[arr_cls == nod] = np.nan
            arr_cls = np.where(arr_cls == nod, 0, arr_cls).astype(np.int32)

        class_stack.append(arr_cls)
        conf_stack.append(arr_conf)

    class_stack = np.stack(class_stack, axis=0)  # (n_tracks, rows, cols)
    conf_stack  = np.stack(conf_stack, axis=0)

    # --- build highest-confidence mosaic -----------------------------------
    conf_stack[np.isnan(conf_stack)] = -np.inf
    idx   = np.argmax(conf_stack, axis=0)               # (rows, cols)
    final = np.take_along_axis(class_stack, idx[None,:,:], axis=0)[0]
    final[np.all(np.isneginf(conf_stack), axis=0)] = 0

    # --- save mosaic --------------------------------------------------------
    base_tr, base_country, _, _ = tracks[0]
    out_dir = base_dir / base_tr / 'classification_results'
    out_tif = out_dir / f"{base_country}_final_classification.tif"
    drv     = gdal.GetDriverByName('GTiff')
    ds_out  = drv.Create(str(out_tif), cols, rows, 1, gdal.GDT_Int32)
    ds_out.SetGeoTransform(gt_global)
    ds_out.SetProjection(proj)
    band = ds_out.GetRasterBand(1)
    band.WriteArray(final)
    band.SetNoDataValue(0)
    ds_out.FlushCache()
    print(f"Merged classification saved: {out_tif}")

    # --- compute metrics & areas --------------------------------------------
    ctrl_shp = out_dir / 'samples' / 'control.shp'
    ctrl     = gpd.read_file(str(ctrl_shp))
    inv      = gdal.InvGeoTransform(gt_global)

    true_vals, pred_vals = [], []
    for _, row in ctrl.iterrows():
        px = int(inv[0] + inv[1]*row.geometry.x + inv[2]*row.geometry.y)
        py = int(inv[3] + inv[4]*row.geometry.x + inv[5]*row.geometry.y)
        if 0 <= px < cols and 0 <= py < rows:
            t = int(row['crop_id'])
            p = int(final[py, px])
            if t > 0 and p > 0:
                true_vals.append(t)
                pred_vals.append(p)

    labels = sorted(set(true_vals + pred_vals))
    cm     = confusion_matrix(true_vals, pred_vals, labels=labels)
    prec, rec, f1, _ = precision_recall_fscore_support(
        true_vals, pred_vals,
        labels=labels,
        average=None,
        zero_division=0
    )
    total = cm.sum()
    exp   = (cm.sum(axis=0) * cm.sum(axis=1)).sum() / (total**2)
    oa    = np.trace(cm) / total
    kappa = (oa - exp) / (1 - exp) if (1 - exp) else np.nan

    resx, resy = abs(gt_global[1]), abs(gt_global[5])
    area_ha    = resx * resy / 10000
    areas      = [{
        'Class':   c,
        'Area_ha': round((final == c).sum() * area_ha, 2)
    } for c in labels]

    # --- write Excel report -------------------------------------------------
    xlsx = out_dir / f"{base_country}_final_metrics.xlsx"
    wb   = openpyxl.Workbook()
    sh   = wb.active
    sh.title = 'Results'

    # Confusion matrix table
    sh.cell(1,1,'Confusion Matrix').font = Font(bold=True)
    for j, lbl in enumerate(labels, start=2):
        sh.cell(2,j,lbl).font = Font(bold=True)
    for i, lbl in enumerate(labels, start=3):
        sh.cell(i,1,lbl).font = Font(bold=True)
        for j in range(len(labels)):
            sh.cell(i,j+2,int(cm[i-3,j]))

    # Overall accuracy & kappa
    r0 = 3 + len(labels)
    sh.cell(r0,1,'Overall Accuracy').font = Font(bold=True)
    sh.cell(r0,2,round(oa,2))
    sh.cell(r0+1,1,'Kappa').font          = Font(bold=True)
    sh.cell(r0+1,2,round(kappa,2))

    # Per‐class recall/precision/F1
    r1 = r0 + 3
    hdrs = ['Class','Producer Acc','User Acc','F1-score']
    for j, h in enumerate(hdrs, start=1):
        sh.cell(r1,j,h).font = Font(bold=True)
    for idx, c in enumerate(labels, start=r1+1):
        sh.cell(idx,1,c)
        sh.cell(idx,2,round(rec[idx-r1-1],2))
        sh.cell(idx,3,round(prec[idx-r1-1],2))
        sh.cell(idx,4,round(f1[idx-r1-1],2))

    # Area per class
    ra = r1 + 1 + len(labels) + 1
    sh.cell(ra,1,'Areas (ha)').font = Font(bold=True)
    sh.cell(ra+1,1,'Class').font   = Font(bold=True)
    sh.cell(ra+1,2,'Area_ha').font = Font(bold=True)
    for i, a in enumerate(areas, start=ra+2):
        sh.cell(i,1,a['Class'])
        sh.cell(i,2,a['Area_ha'])

    wb.save(str(xlsx))
    print(f"Final metrics saved: {xlsx}")

if __name__ == '__main__':
    main()
