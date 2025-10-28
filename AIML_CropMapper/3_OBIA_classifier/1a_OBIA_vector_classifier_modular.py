import os
import argparse
from pathlib import Path
import subprocess
import sys
import geopandas as gpd
import numpy as np
import pandas as pd
from osgeo import gdal, ogr, osr
from sklearn.metrics import confusion_matrix, precision_recall_fscore_support
import openpyxl
from openpyxl.styles import Font

# --- Configuration (Global) ---

# Configuration paths
base_dir = Path("D:/AIML_CropMapper/workingDir")
aux_dir = Path("D:/AIML_CropMapper/auxiliary_files")
track_regions = {
    'P1': 'AU', 'P1a': 'AU',
    'P2': 'IR', 'P3': 'NL',
    'P4': 'PT', 'P4a': 'PT'
}
TOTAL_STAGES = 11

# --- Main Pipeline Class ---

class ProcessingPipeline:
    def __init__(self, track):
        self.track = track
        self.country = track_regions[track]
        self.total_stages = TOTAL_STAGES
        print(f"Initializing pipeline for Track: {self.track}, Country: {self.country}")

        # --- 1. Define all paths ---
        self.base_dir = base_dir
        self.aux_dir = aux_dir
        self.proc_dir = self.base_dir / self.track / 'processed_raster'
        self.out_dir = self.base_dir / self.track / 'classification_results'
        self.samples_dir = self.out_dir / 'samples'
        self.model_dir = self.out_dir / 'train_model'
        self.seg_dir = self.out_dir / 'segmentation'
        self.class_dir = self.out_dir / 'classification'
        
        # Create dirs
        for d in [self.samples_dir, self.model_dir, self.seg_dir, self.class_dir]: 
            d.mkdir(parents=True, exist_ok=True)

        # --- 2. Resolve input raster ---
        self.hdr = next(self.proc_dir.glob(f"*_{self.track}_*_VH_VV.hdr"), None)
        if not self.hdr:
            raise FileNotFoundError(f"No HDR file found for track {self.track} in {self.proc_dir}")
        self.ras = self._resolve_raster(self.hdr)
        print(f"Input raster found: {self.ras}")

        # --- 3. Define all output file paths ---
        self.seg_shp = self.seg_dir / f"{self.country}_{self.track}_segmentation.shp"
        self.sample_shp = self.aux_dir / 'shapefiles_samples' / f"{self.country}_{self.track}" / 'samples.shp'
        self.learn_shp = self.samples_dir / 'learn.shp'
        self.control_shp = self.samples_dir / 'control.shp'
        self.sel_shp = self.samples_dir / f"{self.country}_{self.track}_learn_selected.shp"
        self.class_shp = self.class_dir / f"{self.country}_{self.track}_classified.shp"
        self.class_tif = self.class_dir / f"{self.country}_{self.track}_classified.tif"
        self.conf_tif = self.class_dir / f"{self.country}_{self.track}_confidence_map.tif"
        self.cutline_shp = self.proc_dir / f"{self.country}_{self.track}_valid_coverage.shp"
        self.masked_class = self.class_dir / f"{self.country}_{self.track}_classified_masked.tif"
        self.masked_conf = self.class_dir / f"{self.country}_{self.track}_confidence_masked.tif"
        self.metrics_fp = self.class_dir / f"{self.country}_{self.track}_metrics.xlsx"
        
        # --- 4. Define default parameters for configurable stages ---
        self.stage1_params = {
            'spatialr': 15, 'ranger': 6, 'minsize': 50,
            'tilesizex': 8192, 'tilesizey': 8192, 'ram': 100128
        }
        self.stage2_params = {
            'learn_frac': 0.7, 'random_state': 42
        }
        self.stage4_params = {
            'classifier': 'rf',
            'rf_max': 110, 'rf_min': 2, 'rf_var': 16, 'rf_cat': 16, 'rf_acc': 0.01,
            'svm_c': 1.0, 'svm_k': 'linear' # Example for other classifiers
        }
        
        # --- 5. Shared state variables ---
        self.feat_str = "" # Will be set in stage 4 and used in stage 5


    # --- Utility Methods ---

    def _run_cmd(self, cmd, stage, desc):
        print(f"[Stage {stage}/{self.total_stages}] {desc}")
        proc = subprocess.Popen(cmd, shell=True, stdout=sys.stdout, stderr=sys.stderr)
        proc.communicate()
        if proc.returncode != 0:
            raise RuntimeError(f"Stage {stage} failed: {cmd}")
        print(f"Completed stage {stage}/{self.total_stages}\n")

    def _resolve_raster(self, hdr):
        for ext in ['.tif', '.img', '.hdr']:
            p = hdr.with_suffix(ext)
            if p.exists(): return p
        raise FileNotFoundError(f"No raster for {hdr.stem}")

    def _raster_to_cutline(self, input_tif, cutline_shp, stage):
        print(f"[Stage {stage}/{self.total_stages}] Generating cutline from raster valid pixels")
        ds = gdal.Open(str(input_tif))
        band = ds.GetRasterBand(1)
        arr = band.ReadAsArray()
        nodata = band.GetNoDataValue() if band.GetNoDataValue() is not None else 0

        drv_mem = gdal.GetDriverByName('MEM')
        mask_ds = drv_mem.Create('', ds.RasterXSize, ds.RasterYSize, 1, gdal.GDT_Byte)
        mask_ds.SetGeoTransform(ds.GetGeoTransform())
        mask_ds.SetProjection(ds.GetProjection())
        mask_band = mask_ds.GetRasterBand(1)
        mask_arr = (arr != nodata).astype(np.uint8)
        mask_band.WriteArray(mask_arr)
        mask_band.SetNoDataValue(0)

        shp_drv = ogr.GetDriverByName('ESRI Shapefile')
        if os.path.exists(str(cutline_shp)): shp_drv.DeleteDataSource(str(cutline_shp))
        out_ds = shp_drv.CreateDataSource(str(cutline_shp))
        srs = osr.SpatialReference(); srs.ImportFromWkt(ds.GetProjection())
        layer = out_ds.CreateLayer('cutline', srs=srs, geom_type=ogr.wkbPolygon)
        layer.CreateField(ogr.FieldDefn('DN', ogr.OFTInteger))
        gdal.Polygonize(mask_band, None, layer, 0, [], callback=None)
        out_ds.Destroy()

        gdf = gpd.read_file(str(cutline_shp))
        valid = gdf[gdf['DN'] == 1]
        if valid.empty:
            raise RuntimeError("No valid coverage in raster.")
        dissolved = valid.dissolve(by=None)
        dissolved['DN'] = 1
        dissolved.to_file(str(cutline_shp))
        print(f"Cutline saved to {cutline_shp}\n")

    def _clip_and_mask(self, input_tif, mask_tif, cutline_shp, out_tif, stage):
        print(f"[Stage {stage}/{self.total_stages}] Clipping to cutline and applying arable mask")
        ds_clip = gdal.Warp('', str(input_tif), format='MEM', cutlineDSName=str(cutline_shp), cropToCutline=True, dstNodata=-9999)
        gt = ds_clip.GetGeoTransform(); proj = ds_clip.GetProjection()
        cols, rows = ds_clip.RasterXSize, ds_clip.RasterYSize
        arr = ds_clip.GetRasterBand(1).ReadAsArray()
        nd = ds_clip.GetRasterBand(1).GetNoDataValue() if ds_clip.GetRasterBand(1).GetNoDataValue() is not None else -9999

        minx, maxy = gt[0], gt[3]
        maxx = minx + cols * gt[1]; miny = maxy + rows * gt[5]
        mem_m = gdal.Warp('', str(mask_tif), format='MEM', width=cols, height=rows,
                          outputBounds=(minx, miny, maxx, miny + rows * abs(gt[5])),
                          dstSRS=proj, resampleAlg=gdal.GRA_NearestNeighbour,
                          srcNodata=0, dstNodata=0)
        mask_arr = mem_m.GetRasterBand(1).ReadAsArray()

        out_arr = np.where((arr != nd) & (mask_arr == 1), arr, nd).astype(np.float32)
        driver = gdal.GetDriverByName('GTiff')
        out_ds = driver.Create(str(out_tif), cols, rows, 1, gdal.GDT_Float32)
        out_ds.SetGeoTransform(gt); out_ds.SetProjection(proj)
        bnd = out_ds.GetRasterBand(1); bnd.WriteArray(out_arr); bnd.SetNoDataValue(nd)
        out_ds.FlushCache()
        print(f"Completed stage {stage}/{self.total_stages}\n")


    # --- Stage 1: Segmentation ---
    def stage_1_segmentation(self, **kwargs):
        # Use kwargs to update defaults
        params = self.stage1_params.copy()
        params.update(kwargs)
        stage = 1
        
        if not self.seg_shp.exists():
            cmd = (
                f"otbcli_LargeScaleMeanShift -in {self.ras} -spatialr {params['spatialr']} "
                f"-ranger {params['ranger']} -minsize {params['minsize']} "
                f"-tilesizex {params['tilesizex']} -tilesizey {params['tilesizey']} "
                f"-mode vector -mode.vector.out {self.seg_shp} "
                f"-cleanup false -ram {params['ram']}"
            )
            self._run_cmd(cmd, stage, 'Image segmentation')
        else:
            print(f"[Stage {stage}/{self.total_stages}] Segmentation exists, skipping\n")

    # --- Stage 2: Sample Split ---
    def stage_2_split_samples(self, **kwargs):
        params = self.stage2_params.copy()
        params.update(kwargs)
        stage = 2
        
        print(f"[Stage {stage}/{self.total_stages}] Splitting samples ({params['learn_frac']*100:.0f}/{(1-params['learn_frac'])*100:.0f})")
        if not self.sample_shp.exists():
            print(f"ERROR: Input sample file not found: {self.sample_shp}")
            return
            
        self.samples_dir.mkdir(parents=True, exist_ok=True)
        gdf = gpd.read_file(str(self.sample_shp))
        learn = gdf.sample(frac=params['learn_frac'], random_state=params['random_state'])
        control = gdf.drop(learn.index)
        
        learn.to_file(str(self.learn_shp))
        control.to_file(str(self.control_shp))
        print(f"Completed stage {stage}. Total {len(gdf)}, Learn {len(learn)}, Control {len(control)}\n")

    # --- Stage 3: Selection ---
    def stage_3_selection(self):
        stage = 3
        if not self.sel_shp.exists():
            print(f"[Stage {stage}/{self.total_stages}] Running sample selection")
            if not self.learn_shp.exists():
                print("ERROR: Learn samples not found. Run Stage 2 first.")
                return
            if not self.seg_shp.exists():
                print("ERROR: Segmentation not found. Run Stage 1 first.")
                return
                
            pts = gpd.read_file(self.learn_shp)
            polys = gpd.read_file(self.seg_shp)
            if polys.crs != pts.crs: 
                print("Warning: CRS mismatch. Re-projecting segmentation to sample CRS.")
                polys = polys.to_crs(pts.crs)
            
            sel = gpd.sjoin(polys, pts, how='inner', predicate='intersects')
            sel.to_file(self.sel_shp)
            print(f"[Stage {stage}/{self.total_stages}] Selected {len(sel)} features\n")
        else:
            print(f"[Stage {stage}/{self.total_stages}] Selection exists, skipping\n")

    # --- Stage 4: Train Classifier ---
    def stage_4_train_classifier(self, **kwargs):
        params = self.stage4_params.copy()
        params.update(kwargs)
        stage = 4

        if not self.sel_shp.exists():
            print("ERROR: Selected samples file not found. Run Stage 3 first.")
            return

        df_sel = gpd.read_file(self.sel_shp)
        feats = [c for c in df_sel.columns if c.startswith('meanB')]
        self.feat_str = ' '.join(feats) # Store for stage 5
        
        # Handle model name based on classifier
        clf_name = params['classifier']
        model_fn = self.model_dir / f"{self.country}_{self.track}_model.{clf_name}"
        
        if not model_fn.exists() or os.path.getsize(model_fn) == 0:
            if model_fn.exists(): print(f"[Stage {stage}/{self.total_stages}] Empty model, retrain\n")
            
            # Build classifier string
            clf_str = f"-classifier {clf_name}"
            if clf_name == 'rf':
                clf_str += (
                    f" -classifier.rf.max {params['rf_max']} -classifier.rf.min {params['rf_min']} "
                    f" -classifier.rf.var {params['rf_var']} -classifier.rf.cat {params['rf_cat']} "
                    f" -classifier.rf.acc {params['rf_acc']}"
                )
            elif clf_name == 'svm':
                # Example for SVM
                clf_str += f" -classifier.svm.c {params.get('svm_c', 1.0)}"
                clf_str += f" -classifier.svm.k {params.get('svm_k', 'linear')}"
                print(f"Using SVM with params: {clf_str}")
            # ... add elif for other otb classifiers as needed
            
            cmd = (
                f"otbcli_TrainVectorClassifier -io.vd {self.sel_shp} -io.out {model_fn} "
                f"-feat \"{self.feat_str}\" -cfield crop_id {clf_str}"
            )
            self._run_cmd(cmd, stage, f'Train {clf_name.upper()}')
        else:
            print(f"[Stage {stage}/{self.total_stages}] Model exists, skipping\n")
    
    # --- Stage 5: Classification ---
    def stage_5_classify_vector(self):
        stage = 5
        
        # Find the model file based on current params
        clf_name = self.stage4_params['classifier']
        model_file = self.model_dir / f"{self.country}_{self.track}_model.{clf_name}"
        if not model_file.exists():
             print(f"ERROR: Model file {model_file} not found. Run Stage 4 first.")
             return
        if not self.seg_shp.exists():
             print(f"ERROR: Segmentation file {self.seg_shp} not found. Run Stage 1 first.")
             return
             
        # Get feature string. Try to use from stage 4, otherwise regenerate.
        if not self.feat_str:
            print("Feature string not set. Trying to read from selected samples...")
            try:
                if not self.sel_shp.exists():
                     raise FileNotFoundError("sel_shp not found, run stage 3")
                df_sel = gpd.read_file(self.sel_shp)
                feats = [c for c in df_sel.columns if c.startswith('meanB')]
                self.feat_str = ' '.join(feats)
                if not self.feat_str:
                    raise ValueError("No 'meanB' features found in sel_shp")
            except Exception as e:
                print(f"ERROR: Could not determine features: {e}. Run Stage 4 first.")
                return
        
        if not self.class_shp.exists():
            cmd = (
                f"otbcli_VectorClassifier -in {self.seg_shp} -out {self.class_shp} "
                f"-model {model_file} -feat \"{self.feat_str}\" -cfield predicted -confmap true"
            )
            self._run_cmd(cmd, stage, 'Vec class + conf')
        else:
            print(f"[Stage {stage}/{self.total_stages}] Classification exists, skipping\n")

    # --- Stage 6: Rasterize Class ---
    def stage_6_rasterize_class(self):
        stage = 6
        if not self.class_shp.exists():
            print(f"ERROR: Classified shapefile {self.class_shp} not found. Run Stage 5 first.")
            return
        if not self.class_tif.exists():
            cmd = (
                f"otbcli_Rasterization -in {self.class_shp} -out {self.class_tif} "
                f"-mode attribute -mode.attribute.field predicted -spx 10 -spy 10"
            )
            self._run_cmd(cmd, stage, 'Rasterize classification')
        else:
            print(f"[Stage {stage}/{self.total_stages}] Classified TIFF exists, skipping\n")

    # --- Stage 7: Rasterize Confidence ---
    def stage_7_rasterize_confidence(self):
        stage = 7
        if not self.class_shp.exists():
            print(f"ERROR: Classified shapefile {self.class_shp} not found. Run Stage 5 first.")
            return
        if not self.conf_tif.exists():
            cmd = (
                f"otbcli_Rasterization -in {self.class_shp} -out {self.conf_tif} "
                f"-mode attribute -mode.attribute.field confidence -spx 10 -spy 10"
            )
            self._run_cmd(cmd, stage, 'Rasterize confidence')
        else:
            print(f"[Stage {stage}/{self.total_stages}] Confidence TIFF exists, skipping\n")

    # --- Stage 8: Cutline ---
    def stage_8_create_cutline(self):
        stage = 8
        if not self.cutline_shp.exists():
            self._raster_to_cutline(self.ras, self.cutline_shp, stage)
        else:
            print(f"[Stage {stage}/{self.total_stages}] Cutline exists, skipping\n")
    
    # --- Stage 9: Mask Class ---
    def stage_9_mask_class(self):
        stage = 9
        mask_file = self.aux_dir / 'raster_files' / 'EU_arable_areas_mask_3857.tif'
        if not self.class_tif.exists():
            print(f"ERROR: Classified TIF {self.class_tif} not found. Run Stage 6 first.")
            return
        if not self.cutline_shp.exists():
            print(f"ERROR: Cutline {self.cutline_shp} not found. Run Stage 8 first.")
            return
        if not mask_file.exists():
            print(f"ERROR: Arable mask {mask_file} not found.")
            return
            
        if not self.masked_class.exists():
            self._clip_and_mask(self.class_tif, mask_file, self.cutline_shp, self.masked_class, stage)
        else:
            print(f"[Stage {stage}/{self.total_stages}] Masked classification exists, skipping\n")

    # --- Stage 10: Mask Confidence ---
    def stage_10_mask_confidence(self):
        stage = 10
        mask_file = self.aux_dir / 'raster_files' / 'EU_arable_areas_mask_3857.tif'
        if not self.conf_tif.exists():
            print(f"ERROR: Confidence TIF {self.conf_tif} not found. Run Stage 7 first.")
            return
        if not self.cutline_shp.exists():
            print(f"ERROR: Cutline {self.cutline_shp} not found. Run Stage 8 first.")
            return
        if not mask_file.exists():
            print(f"ERROR: Arable mask {mask_file} not found.")
            return
            
        if not self.masked_conf.exists():
            self._clip_and_mask(self.conf_tif, mask_file, self.cutline_shp, self.masked_conf, stage)
        else:
            print(f"[Stage {stage}/{self.total_stages}] Masked confidence exists, skipping\n")
    
    # --- Stage 11: Metrics ---
    def stage_11_calculate_metrics(self):
        stage = 11
        if not self.metrics_fp.exists():
            print(f"[Stage {stage}/{self.total_stages}] Computing metrics and exporting Excel")
            
            if not self.control_shp.exists():
                print(f"ERROR: Control shapefile {self.control_shp} not found. Run Stage 2.")
                return
            if not self.masked_class.exists():
                print(f"ERROR: Masked classification {self.masked_class} not found. Run Stage 9.")
                return
                
            # Read control points and masked classified raster
            ctrl = gpd.read_file(str(self.control_shp))
            ds = gdal.Open(str(self.masked_class))
            arr = ds.GetRasterBand(1).ReadAsArray()
            gt = ds.GetGeoTransform()
            inv = gdal.InvGeoTransform(gt)
            true_vals, pred_vals = [], []
            
            for _, row in ctrl.iterrows():
                try:
                    px = int(inv[0] + inv[1]*row.geometry.x + inv[2]*row.geometry.y)
                    py = int(inv[3] + inv[4]*row.geometry.x + inv[5]*row.geometry.y)
                    if 0 <= px < arr.shape[1] and 0 <= py < arr.shape[0]:
                        t = int(row['crop_id'])
                        p = int(arr[py, px])
                        if t > 0 and p > 0:
                            true_vals.append(t)
                            pred_vals.append(p)
                except Exception as e:
                    print(f"Warning: could not process point {row.geometry}: {e}")
                    
            if not true_vals or not pred_vals:
                print("ERROR: No valid matching true/predicted values found. Cannot compute metrics.")
                return
                
            labels = sorted(list(set(true_vals + pred_vals)))
            cm = confusion_matrix(true_vals, pred_vals, labels=labels)
            precisions, recalls, f1s, _ = precision_recall_fscore_support(true_vals, pred_vals, labels=labels, average=None, zero_division=0)
            wp, wr, wf, _ = precision_recall_fscore_support(true_vals, pred_vals, average='weighted', zero_division=0)
            
            total = np.sum(cm)
            oa = np.trace(cm) / total
            sum_po = oa
            sum_pe = np.sum(np.sum(cm, axis=0) * np.sum(cm, axis=1)) / (total**2)
            kappa = (sum_po - sum_pe) / (1 - sum_pe) if (1 - sum_pe) != 0 else np.nan

            # compute area per class
            resx, resy = abs(gt[1]), abs(gt[5])
            area_ha = resx * resy / 10000
            unique_classes, counts = np.unique(arr[arr > 0], return_counts=True)
            class_areas = dict(zip(unique_classes, counts))
            
            areas = [{'Class': c, 'Area_ha': round(class_areas.get(c, 0) * area_ha, 2)} for c in labels]
            
            # Write Excel
            wb = openpyxl.Workbook()
            sh = wb.active; sh.title = 'Results'
            
            # Confusion
            sh.cell(row=1, column=1, value='Confusion Matrix').font = Font(bold=True)
            sh.cell(row=2, column=1, value='True \\ Pred').font = Font(bold=True)
            for j, lbl in enumerate(labels, start=2): 
                sh.cell(row=2, column=j, value=lbl).font = Font(bold=True)
            for i, lbl in enumerate(labels, start=3):
                sh.cell(row=i, column=1, value=lbl).font = Font(bold=True)
                for j, _ in enumerate(labels):
                    sh.cell(row=i, column=j+2, value=int(cm[i-3, j]))
            
            # OA & Kappa
            base = 4 + len(labels)
            sh.cell(row=base, column=1, value='Overall Accuracy').font = Font(bold=True)
            sh.cell(row=base, column=2, value=round(oa, 4))
            sh.cell(row=base+1, column=1, value='Kappa').font = Font(bold=True)
            sh.cell(row=base+1, column=2, value=round(kappa, 4))
            
            # Classification metrics
            start = base+3
            headers = ['Class','Producer Acc (Recall)','User Acc (Precision)','F1-score']
            for j, h in enumerate(headers, start=1): 
                sh.cell(row=start, column=j, value=h).font = Font(bold=True)
            for idx, c in enumerate(labels):
                row_idx = start + 1 + idx
                sh.cell(row=row_idx, column=1, value=c)
                sh.cell(row=row_idx, column=2, value=round(recalls[idx], 4))
                sh.cell(row=row_idx, column=3, value=round(precisions[idx], 4))
                sh.cell(row=row_idx, column=4, value=round(f1s[idx], 4))
            
            # Areas
            ar0 = start + 1 + len(labels) + 1
            sh.cell(row=ar0, column=1, value='Areas (ha)').font = Font(bold=True)
            sh.cell(row=ar0+1, column=1, value='Class').font = Font(bold=True)
            sh.cell(row=ar0+1, column=2, value='Area_ha').font = Font(bold=True)
            for idx, a in enumerate(areas, start=ar0+2):
                sh.cell(row=idx, column=1, value=a['Class'])
                sh.cell(row=idx, column=2, value=a['Area_ha'])
                
            wb.save(str(self.metrics_fp))
            print(f"Metrics and Excel saved to {self.metrics_fp}\n")
        else:
            print(f"[Stage 11/{self.total_stages}] Metrics Excel exists, skipping")
        print(f"All done! Metrics available at {self.metrics_fp}")


# --- Interactive Menu Helpers ---

def get_params(param_dict):
    """Helper to interactively update a parameter dictionary."""
    new_params = param_dict.copy()
    print("--- Current Parameters ---")
    for key, val in new_params.items():
        print(f"  {key}: {val}")
    
    if input("Change parameters? (y/n) [n]: ").lower() != 'y':
        return new_params

    for key, val in new_params.items():
        new_val_str = input(f"Enter new value for '{key}' [{val}]: ")
        if not new_val_str:
            continue # Keep default
        
        # Try to cast to original type
        try:
            original_type = type(val)
            new_params[key] = original_type(new_val_str)
        except ValueError:
            print(f"Invalid value. Keeping default {val}.")
    
    print("--- Updated Parameters ---")
    for key, val in new_params.items():
        print(f"  {key}: {val}")
    return new_params

def get_classifier_params(param_dict):
    """Special helper for classifier params."""
    new_params = param_dict.copy()
    print("--- Current Parameters ---")
    for key, val in new_params.items():
        print(f"  {key}: {val}")
    
    if input("Change parameters? (y/n) [n]: ").lower() != 'y':
        return new_params

    # 1. Choose classifier
    clf = input(f"Enter classifier (e.g., rf, svm) [{new_params['classifier']}]: ") or new_params['classifier']
    new_params['classifier'] = clf.lower()

    # 2. Update params for that classifier
    print(f"\n--- Setting parameters for {clf.upper()} ---")
    if clf == 'rf':
        prefix = 'rf_'
    elif clf == 'svm':
        prefix = 'svm_'
    else:
        print(f"No specific parameters defined for '{clf}'. No parameters will be updated.")
        prefix = None
        
    if prefix:
        for key in [k for k in new_params if k.startswith(prefix)]:
            val = new_params[key]
            new_val_str = input(f"Enter new value for '{key}' [{val}]: ")
            if new_val_str:
                try:
                    new_params[key] = type(val)(new_val_str)
                except ValueError:
                    print(f"Invalid value. Keeping default {val}.")

    print("--- Updated Parameters ---")
    for key, val in new_params.items():
        print(f"  {key}: {val}")
    return new_params

# --- Main Execution ---

def main_menu(pipeline):
    """Displays the main interactive menu."""
    
    menu = f"""
    --- Processing Pipeline Menu ---
    Track: {pipeline.track} ({pipeline.country})

    [1] Stage 1: Segmentation
    [2] Stage 2: Split Samples (current: {pipeline.stage2_params['learn_frac']*100:.0f}%)
    [3] Stage 3: Select Training Polygons
    [4] Stage 4: Train Classifier (current: {pipeline.stage4_params['classifier'].upper()})
    [5] Stage 5: Classify Vector
    [6] Stage 6: Rasterize Classification
    [7] Stage 7: Rasterize Confidence
    [8] Stage 8: Create Valid-Pixel Cutline
    [9] Stage 9: Mask Classification
    [10] Stage 10: Mask Confidence
    [11] Stage 11: Calculate Metrics

    [A] Run All Stages (using current parameters)
    [Q] Quit

    Enter your choice: 
    """

    while True:
        choice = input(menu).strip().upper()
        
        try:
            if choice == '1':
                print("\n--- STAGE 1: SEGMENTATION ---")
                new_params = get_params(pipeline.stage1_params)
                pipeline.stage1_params.update(new_params)
                pipeline.stage_1_segmentation(**pipeline.stage1_params)
            
            elif choice == '2':
                print("\n--- STAGE 2: SPLIT SAMPLES ---")
                new_params = get_params(pipeline.stage2_params)
                pipeline.stage2_params.update(new_params)
                pipeline.stage_2_split_samples(**pipeline.stage2_params)
            
            elif choice == '3':
                print("\n--- STAGE 3: SELECTION ---")
                pipeline.stage_3_selection()
            
            elif choice == '4':
                print("\n--- STAGE 4: TRAIN CLASSIFIER ---")
                new_params = get_classifier_params(pipeline.stage4_params)
                pipeline.stage4_params.update(new_params)
                pipeline.stage_4_train_classifier(**pipeline.stage4_params)
            
            elif choice == '5':
                print("\n--- STAGE 5: CLASSIFY VECTOR ---")
                pipeline.stage_5_classify_vector()
            
            elif choice == '6':
                print("\n--- STAGE 6: RASTERIZE CLASS ---")
                pipeline.stage_6_rasterize_class()
                
            elif choice == '7':
                print("\n--- STAGE 7: RASTERIZE CONFIDENCE ---")
                pipeline.stage_7_rasterize_confidence()

            elif choice == '8':
                print("\n--- STAGE 8: CREATE CUTLINE ---")
                pipeline.stage_8_create_cutline()

            elif choice == '9':
                print("\n--- STAGE 9: MASK CLASS ---")
                pipeline.stage_9_mask_class()

            elif choice == '10':
                print("\n--- STAGE 10: MASK CONFIDENCE ---")
                pipeline.stage_10_mask_confidence()

            elif choice == '11':
                print("\n--- STAGE 11: CALCULATE METRICS ---")
                pipeline.stage_11_calculate_metrics()
            
            elif choice == 'A':
                print("\n--- RUNNING ALL STAGES ---")
                print("(Using currently configured parameters)")
                
                print("\n--- STAGE 1: SEGMENTATION ---")
                pipeline.stage_1_segmentation(**pipeline.stage1_params)
                print("\n--- STAGE 2: SPLIT SAMPLES ---")
                pipeline.stage_2_split_samples(**pipeline.stage2_params)
                print("\n--- STAGE 3: SELECTION ---")
                pipeline.stage_3_selection()
                print("\n--- STAGE 4: TRAIN CLASSIFIER ---")
                pipeline.stage_4_train_classifier(**pipeline.stage4_params)
                print("\n--- STAGE 5: CLASSIFY VECTOR ---")
                pipeline.stage_5_classify_vector()
                print("\n--- STAGE 6: RASTERIZE CLASS ---")
                pipeline.stage_6_rasterize_class()
                print("\n--- STAGE 7: RASTERIZE CONFIDENCE ---")
                pipeline.stage_7_rasterize_confidence()
                print("\n--- STAGE 8: CREATE CUTLINE ---")
                pipeline.stage_8_create_cutline()
                print("\n--- STAGE 9: MASK CLASS ---")
                pipeline.stage_9_mask_class()
                print("\n--- STAGE 10: MASK CONFIDENCE ---")
                pipeline.stage_10_mask_confidence()
                print("\n--- STAGE 11: CALCULATE METRICS ---")
                pipeline.stage_11_calculate_metrics()
                print("\n--- ALL STAGES COMPLETE ---")
            
            elif choice == 'Q':
                print("Exiting.")
                break
            
            else:
                print("Invalid choice. Please try again.")

        except Exception as e:
            print(f"\n---!!! AN ERROR OCCURRED !!!---")
            print(f"Error: {e}")
            print("Please check your inputs, paths, and OTB installation.")
            print("Returning to main menu.\n")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Modular OBIA Classification Pipeline")
    parser.add_argument('--track', required=True, help="Processing track ID (e.g., P1, P2)")
    args = parser.parse_args()
    
    try:
        # Initialize the pipeline class
        pipeline = ProcessingPipeline(track=args.track)
        # Run the interactive menu
        main_menu(pipeline)
    except Exception as e:
        print(f"\n---!!! A FATAL ERROR OCCURRED ON INITIALIZATION !!!---")
        print(f"Error: {e}")
        print("Please check your configuration, paths, and track ID.")
        sys.exit(1)