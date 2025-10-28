import os
import argparse
from pathlib import Path
import subprocess
import sys
import geopandas as gpd
import numpy as np
import pandas as pd
from osgeo import gdal, ogr, osr
from sklearn.metrics import confusion_matrix, precision_recall_fscore_support

# Configuration paths
base_dir = Path("D:/AIML_CropMapper/workingDir")
aux_dir = Path("D:/AIML_CropMapper/auxiliary_files")
track_regions = {
    'P1': 'AU', 'P1a': 'AU',
    'P2': 'IR', 'P3': 'NL',
    'P4': 'PT', 'P4a': 'PT'
}
TOTAL_STAGES = 11

# Utility: run shell command
def run_cmd(cmd, stage, desc):
    print(f"[Stage {stage}/{TOTAL_STAGES}] {desc}")
    proc = subprocess.Popen(cmd, shell=True, stdout=sys.stdout, stderr=sys.stderr)
    proc.communicate()
    if proc.returncode != 0:
        raise RuntimeError(f"Stage {stage} failed: {cmd}")
    print(f"Completed stage {stage}/{TOTAL_STAGES}\n")

# Split samples into learn/control
def split_samples(input_shp, samples_dir, stage):
    print(f"[Stage {stage}/{TOTAL_STAGES}] Splitting samples (70/30)")
    samples_dir.mkdir(parents=True, exist_ok=True)
    gdf = gpd.read_file(str(input_shp))
    learn = gdf.sample(frac=0.7, random_state=42)
    control = gdf.drop(learn.index)
    learn_fp = samples_dir / 'learn.shp'
    control_fp = samples_dir / 'control.shp'
    learn.to_file(str(learn_fp))
    control.to_file(str(control_fp))
    print(f"Completed stage {stage}. Total {len(gdf)}, Learn {len(learn)}, Control {len(control)}\n")
    return learn_fp, control_fp

# Create cutline from raster valid pixels
def raster_to_cutline(input_tif, cutline_shp, stage):
    print(f"[Stage {stage}/{TOTAL_STAGES}] Generating cutline from raster valid pixels")
    ds = gdal.Open(str(input_tif))
    band = ds.GetRasterBand(1)
    arr = band.ReadAsArray()
    nodata = band.GetNoDataValue() if band.GetNoDataValue() is not None else 0

    drv_mem = gdal.GetDriverByName('MEM')
    mask_ds = drv_mem.Create('', ds.RasterXSize, ds.RasterYSize, 1, gdal.GDT_Byte)
    mask_ds.SetGeoTransform(ds.GetGeoTransform())
    mask_ds.SetProjection(ds.GetProjection())
    mask_band = mask_ds.GetRasterBand(1)
    mask_arr = (arr != nodata).astype(np.uint8)
    mask_band.WriteArray(mask_arr)
    mask_band.SetNoDataValue(0)

    shp_drv = ogr.GetDriverByName('ESRI Shapefile')
    if os.path.exists(cutline_shp): shp_drv.DeleteDataSource(str(cutline_shp))
    out_ds = shp_drv.CreateDataSource(str(cutline_shp))
    srs = osr.SpatialReference(); srs.ImportFromWkt(ds.GetProjection())
    layer = out_ds.CreateLayer('cutline', srs=srs, geom_type=ogr.wkbPolygon)
    layer.CreateField(ogr.FieldDefn('DN', ogr.OFTInteger))
    gdal.Polygonize(mask_band, None, layer, 0, [], callback=None)
    out_ds.Destroy()

    gdf = gpd.read_file(str(cutline_shp))
    valid = gdf[gdf['DN'] == 1]
    if valid.empty:
        raise RuntimeError("No valid coverage in raster.")
    dissolved = valid.dissolve(by=None)
    dissolved['DN'] = 1
    dissolved.to_file(str(cutline_shp))
    print(f"Cutline saved to {cutline_shp}\n")

# Clip and mask by cutline and arable mask
def clip_and_mask(input_tif, mask_tif, cutline_shp, out_tif, stage):
    print(f"[Stage {stage}/{TOTAL_STAGES}] Clipping to cutline and applying arable mask")
    ds_clip = gdal.Warp('', str(input_tif), format='MEM', cutlineDSName=str(cutline_shp), cropToCutline=True, dstNodata=-9999)
    gt = ds_clip.GetGeoTransform(); proj = ds_clip.GetProjection()
    cols, rows = ds_clip.RasterXSize, ds_clip.RasterYSize
    arr = ds_clip.GetRasterBand(1).ReadAsArray()
    nd = ds_clip.GetRasterBand(1).GetNoDataValue() if ds_clip.GetRasterBand(1).GetNoDataValue() is not None else -9999

    minx, maxy = gt[0], gt[3]
    maxx = minx + cols * gt[1]; miny = maxy + rows * gt[5]
    mem_m = gdal.Warp('', str(mask_tif), format='MEM', width=cols, height=rows,
                      outputBounds=(minx, miny, maxx, miny + rows * abs(gt[5])),
                      dstSRS=proj, resampleAlg=gdal.GRA_NearestNeighbour,
                      srcNodata=0, dstNodata=0)
    mask_arr = mem_m.GetRasterBand(1).ReadAsArray()

    out_arr = np.where((arr != nd) & (mask_arr == 1), arr, nd).astype(np.float32)
    driver = gdal.GetDriverByName('GTiff')
    out_ds = driver.Create(str(out_tif), cols, rows, 1, gdal.GDT_Float32)
    out_ds.SetGeoTransform(gt); out_ds.SetProjection(proj)
    bnd = out_ds.GetRasterBand(1); bnd.WriteArray(out_arr); bnd.SetNoDataValue(nd)
    out_ds.FlushCache()
    print(f"Completed stage {stage}/{TOTAL_STAGES}\n")

# Resolve raster file from header
def resolve_raster(hdr):
    for ext in ['.tif', '.img', '.hdr']:
        p = hdr.with_suffix(ext)
        if p.exists(): return p
    raise FileNotFoundError(f"No raster for {hdr.stem}")

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--track', required=True)
    args = parser.parse_args()
    track = args.track; country = track_regions[track]

    proc_dir = base_dir / track / 'processed_raster'
    out_dir = base_dir / track / 'classification_results'
    samples_dir = out_dir / 'samples'; model_dir = out_dir / 'train_model'
    seg_dir = out_dir / 'segmentation'; class_dir = out_dir / 'classification'
    for d in [samples_dir, model_dir, seg_dir, class_dir]: d.mkdir(parents=True, exist_ok=True)

    hdr = next(proc_dir.glob(f"*_{track}_*_VH_VV.hdr"), None)
    ras = resolve_raster(hdr)

    # Stage 1: segmentation
    seg_shp = seg_dir / f"{country}_{track}_segmentation.shp"
    if not seg_shp.exists():
        run_cmd(
            f"otbcli_LargeScaleMeanShift -in {ras} -spatialr 15 -ranger 6 -minsize 50 "
            f"-tilesizex 8192 -tilesizey 8192 -mode vector -mode.vector.out {seg_shp} "
            f"-cleanup false -ram 100128", 1, 'Image segmentation'
        )
    else:
        print(f"[Stage 1/{TOTAL_STAGES}] Segmentation exists, skipping\n")

    # Stage 2: sample split
    sample_shp = aux_dir / 'shapefiles_samples' / f"{country}_{track}" / 'samples.shp'
    learn_shp, control_shp = split_samples(sample_shp, samples_dir, 2)

    # Stage 3: selection
    sel_shp = samples_dir / f"{country}_{track}_learn_selected.shp"
    if not sel_shp.exists():
        pts = gpd.read_file(learn_shp); polys = gpd.read_file(seg_shp)
        if polys.crs != pts.crs: polys = polys.to_crs(pts.crs)
        sel = gpd.sjoin(polys, pts, how='inner', op='intersects')
        sel.to_file(sel_shp); print(f"[Stage 3/{TOTAL_STAGES}] Selected {len(sel)} features\n")
    else:
        print(f"[Stage 3/{TOTAL_STAGES}] Selection exists, skipping\n")

    # Stage 4: train RF
    df_sel = gpd.read_file(sel_shp)
    feats = [c for c in df_sel.columns if c.startswith('meanB')]
    feat_str = ' '.join(feats)
    model_fn = model_dir / f"{country}_{track}_model.rf"
    if not model_fn.exists() or os.path.getsize(model_fn) == 0:
        if model_fn.exists(): print(f"[Stage 4/{TOTAL_STAGES}] Empty model, retrain\n")
        run_cmd(
            f"otbcli_TrainVectorClassifier -io.vd {sel_shp} -io.out {model_fn} -feat {feat_str} "
            f"-cfield crop_id -classifier rf -classifier.rf.max 110 -classifier.rf.min 2 "
            f"-classifier.rf.var 16 -classifier.rf.cat 16 -classifier.rf.acc 0.01",
            4, 'Train RF'
        )
    else:
        print(f"[Stage 4/{TOTAL_STAGES}] Model exists, skipping\n")

    # Stage 5: classification + confidence
    class_shp = class_dir / f"{country}_{track}_classified.shp"
    if not class_shp.exists():
        run_cmd(f"otbcli_VectorClassifier -in {seg_shp} -out {class_shp} -model {model_fn} -feat {feat_str} -cfield predicted -confmap true", 5, 'Vec class + conf')
    else:
        print(f"[Stage 5/{TOTAL_STAGES}] Classification exists, skipping\n")

    # Stage 6: rasterize classification
    class_tif = class_dir / f"{country}_{track}_classified.tif"
    if not class_tif.exists():
        run_cmd(f"otbcli_Rasterization -in {class_shp} -out {class_tif} -mode attribute -mode.attribute.field predicted -spx 10 -spy 10", 6, 'Rasterize classification')
    else:
        print(f"[Stage 6/{TOTAL_STAGES}] Classified TIFF exists, skipping\n")

    # Stage 7: rasterize confidence
    conf_tif = class_dir / f"{country}_{track}_confidence_map.tif"
    if not conf_tif.exists():
        run_cmd(f"otbcli_Rasterization -in {class_shp} -out {conf_tif} -mode attribute -mode.attribute.field confidence -spx 10 -spy 10", 7, 'Rasterize confidence')
    else:
        print(f"[Stage 7/{TOTAL_STAGES}] Confidence TIFF exists, skipping\n")

    # Stage 8: cutline extraction
    cutline_shp = proc_dir / f"{country}_{track}_valid_coverage.shp"
    if not cutline_shp.exists():
        raster_to_cutline(ras, cutline_shp, 8)
    else:
        print(f"[Stage 8/{TOTAL_STAGES}] Cutline exists, skipping\n")

    # Stage 9: mask classification
    masked_class = class_dir / f"{country}_{track}_classified_masked.tif"
    if not masked_class.exists():
        clip_and_mask(class_tif, aux_dir / 'raster_files' / 'EU_arable_areas_mask_3857.tif', cutline_shp, masked_class, 9)
    else:
        print(f"[Stage 9/{TOTAL_STAGES}] Masked classification exists, skipping\n")

    # Stage 10: mask confidence
    masked_conf = class_dir / f"{country}_{track}_confidence_masked.tif"
    if not masked_conf.exists():
        clip_and_mask(conf_tif, aux_dir / 'raster_files' / 'EU_arable_areas_mask_3857.tif', cutline_shp, masked_conf, 10)
    else:
        print(f"[Stage 10/{TOTAL_STAGES}] Masked confidence exists, skipping\n")

    # Stage 11: metrics & Excel export
    metrics_fp = class_dir / f"{country}_{track}_metrics.xlsx"
    if not metrics_fp.exists():
        print(f"[Stage 11/{TOTAL_STAGES}] Computing metrics and exporting Excel")
        # Read control points and masked classified raster
        ctrl = gpd.read_file(str(control_shp))
        ds = gdal.Open(str(masked_class))
        arr = ds.GetRasterBand(1).ReadAsArray()
        gt = ds.GetGeoTransform()
        inv = gdal.InvGeoTransform(gt)
        true_vals, pred_vals = [], []
        for _, row in ctrl.iterrows():
            px = int(inv[0] + inv[1]*row.geometry.x + inv[2]*row.geometry.y)
            py = int(inv[3] + inv[4]*row.geometry.x + inv[5]*row.geometry.y)
            if 0 <= px < arr.shape[1] and 0 <= py < arr.shape[0]:
                t = int(row['crop_id'])
                p = int(arr[py, px])
                if t > 0 and p > 0:
                    true_vals.append(t)
                    pred_vals.append(p)
        labels = sorted(set(true_vals + pred_vals))
        cm = confusion_matrix(true_vals, pred_vals, labels=labels)
        precisions, recalls, f1s, _ = precision_recall_fscore_support(true_vals, pred_vals, labels=labels, average=None, zero_division=0)
        wp, wr, wf, _ = precision_recall_fscore_support(true_vals, pred_vals, average='weighted', zero_division=0)
        total = np.sum(cm)
        exp = np.sum(np.sum(cm, axis=0)*np.sum(cm, axis=1)) / (total**2)
        oa = np.trace(cm) / total
        kappa = (oa - exp) / (1 - exp) if (1 - exp) != 0 else np.nan
        # compute area per class
        resx, resy = abs(gt[1]), abs(gt[5])
        area_ha = resx * resy / 10000
        areas = [{'Class': c, 'Area_ha': round(np.sum(arr == c)*area_ha, 2)} for c in labels]
        # Write Excel
        import openpyxl
        from openpyxl.styles import Font
        wb = openpyxl.Workbook()
        sh = wb.active; sh.title = 'Results'
        # Confusion
        sh.cell(row=1, column=1, value='Confusion Matrix').font = Font(bold=True)
        for j, lbl in enumerate(labels, start=2): sh.cell(row=2, column=j, value=lbl).font = Font(bold=True)
        for i, lbl in enumerate(labels, start=3):
            sh.cell(row=i, column=1, value=lbl).font = Font(bold=True)
            for j, _ in enumerate(labels):
                sh.cell(row=i, column=j+2, value=int(cm[i-3, j]))
        # OA & Kappa
        base = 3 + len(labels)
        sh.cell(row=base, column=1, value='Overall Accuracy').font = Font(bold=True)
        sh.cell(row=base, column=2, value=round(oa,2))
        sh.cell(row=base+1, column=1, value='Kappa').font = Font(bold=True)
        sh.cell(row=base+1, column=2, value=round(kappa,2))
        # Classification metrics
        start = base+3
        headers = ['Class','Producer Acc (Recall)','User Acc (Precision)','F1-score']
        for j, h in enumerate(headers, start=1): sh.cell(row=start, column=j, value=h).font = Font(bold=True)
        for idx, c in enumerate(labels, start=start+1):
            sh.cell(row=idx, column=1, value=c)
            sh.cell(row=idx, column=2, value=round(recalls[idx-start-1],2))
            sh.cell(row=idx, column=3, value=round(precisions[idx-start-1],2))
            sh.cell(row=idx, column=4, value=round(f1s[idx-start-1],2))
        # Areas
        ar0 = start+1+len(labels)+1
        sh.cell(row=ar0, column=1, value='Areas (ha)').font = Font(bold=True)
        sh.cell(row=ar0+1, column=1, value='Class').font = Font(bold=True)
        sh.cell(row=ar0+1, column=2, value='Area_ha').font = Font(bold=True)
        for idx, a in enumerate(areas, start=ar0+2):
            sh.cell(row=idx, column=1, value=a['Class'])
            sh.cell(row=idx, column=2, value=a['Area_ha'])
        wb.save(str(metrics_fp))
        print(f"Metrics and Excel saved to {metrics_fp}")
    else:
        print(f"[Stage 11/{TOTAL_STAGES}] Metrics Excel exists, skipping")
    print(f"All done! Metrics available at {metrics_fp}")
