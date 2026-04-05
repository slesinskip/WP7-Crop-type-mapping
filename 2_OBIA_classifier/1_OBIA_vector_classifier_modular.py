import os
import argparse
from pathlib import Path
import subprocess
import sys
import geopandas as gpd
import numpy as np
import pandas as pd
from osgeo import gdal, ogr, osr
from sklearn.metrics import confusion_matrix, precision_recall_fscore_support
import openpyxl
from openpyxl.styles import Font

# --- Configuration (Global) ---

# python 1_OBIA_vector_classifier_modular.py -track P1a

# Base Paths provided by user
base_dir = Path("D:/AIML_CropMapper_Cloud/workingDir")
aux_dir = Path("D:/AIML_CropMapper_Cloud/auxiliary_files")

# OTB Installation Path
otb_dir = Path("D:/AIML_CropMapper_Cloud/2_OBIA_classifier/OTB-6.2.0-Win64")

# Track to Country Mapping
track_regions = {
    'P1': 'AT', 'P1a': 'AT',
    'P2': 'IE', 'P2a': 'IE',
    'P3': 'NL',
    'P4': 'PT', 'P4a': 'PT'
}
TOTAL_STAGES = 11


# --- Main Pipeline Class ---

class ProcessingPipeline:
    def __init__(self, track):
        self.track = track
        try:
            self.country = track_regions[track]
        except KeyError:
            print(f"Error: Track '{track}' not defined in track_regions configuration.")
            sys.exit(1)

        self.total_stages = TOTAL_STAGES
        print(f"Initializing pipeline for Track: {self.track}, Country: {self.country}")

        # --- 1. Define all paths ---
        self.base_dir = base_dir
        self.aux_dir = aux_dir
        self.proc_dir = self.base_dir / self.track / 'processed_raster'
        self.out_dir = self.base_dir / self.track / 'classification_results'
        self.samples_dir = self.out_dir / 'samples'
        self.model_dir = self.out_dir / 'train_model'
        self.seg_dir = self.out_dir / 'segmentation'
        self.class_dir = self.out_dir / 'classification'

        # --- Ensure directories exist immediately upon init ---
        self._ensure_directories()

        # --- 2. Resolve input raster ---
        # Look for files matching various patterns
        search_patterns = [
            f"{self.track}_*_VH_VV*.hdr",  # Your case: P1_2025...
            f"*_{self.track}_*_VH_VV*.hdr",  # Original case: S1_P1...
            f"*{self.track}*.hdr",  # Fallback
        ]

        self.hdr = None
        if self.proc_dir.exists():
            for pattern in search_patterns:
                self.hdr = next(self.proc_dir.glob(pattern), None)
                if self.hdr:
                    break

            if not self.hdr:
                # Check for TIF/IMG without HDR
                tif_search = next(self.proc_dir.glob(f"*{self.track}*.tif"), None)
                if tif_search:
                    self.hdr = tif_search
                else:
                    raise FileNotFoundError(f"No HDR or raster file found for track {self.track} in {self.proc_dir}")

            self.ras = self._resolve_raster(self.hdr)
            print(f"Input raster found: {self.ras}")
        else:
            raise FileNotFoundError(f"Processing directory does not exist: {self.proc_dir}")

        # --- 3. Define all output file paths ---
        self.seg_shp = self.seg_dir / f"{self.country}_{self.track}_segmentation.shp"

        # --- PATH FIX: Smart Search for samples.shp ---
        # We look in 'shapefiles_samples' and try common subfolder names
        samples_base = self.aux_dir / 'shapefiles_samples'

        # Possible locations for samples.shp
        candidate_paths = [
            samples_base / f"{self.country}_{self.track}" / "samples.shp",  # Check shapefiles_samples/AU_P1/samples.shp
            samples_base / self.track / "samples.shp",  # Check shapefiles_samples/P1/samples.shp
            samples_base / self.country / "samples.shp",  # Check shapefiles_samples/AT/samples.shp
            samples_base / "samples.shp"  # Check shapefiles_samples/samples.shp
        ]

        self.sample_shp = None
        for p in candidate_paths:
            if p.exists():
                self.sample_shp = p
                print(f"Training samples found at: {self.sample_shp}")
                break

        if not self.sample_shp:
            print(f"\nCRITICAL WARNING: Could not find 'samples.shp' inside {samples_base}")
            print(f"Checked subfolders: {self.country}_{self.track}, {self.track}, {self.country}")
            # Fallback path
            self.sample_shp = samples_base / f"{self.country}_{self.track}" / "samples.shp"

        # Output paths
        self.learn_shp = self.samples_dir / 'learn.shp'
        self.control_shp = self.samples_dir / 'control.shp'
        self.sel_shp = self.samples_dir / f"{self.country}_{self.track}_learn_selected.shp"
        self.class_shp = self.class_dir / f"{self.country}_{self.track}_classified.shp"
        self.class_tif = self.class_dir / f"{self.country}_{self.track}_classified.tif"
        self.conf_tif = self.class_dir / f"{self.country}_{self.track}_confidence_map.tif"
        self.cutline_shp = self.proc_dir / f"{self.country}_{self.track}_valid_coverage.shp"
        self.masked_class = self.class_dir / f"{self.country}_{self.track}_classified_masked.tif"
        self.masked_conf = self.class_dir / f"{self.country}_{self.track}_confidence_masked.tif"
        self.metrics_fp = self.class_dir / f"{self.country}_{self.track}_metrics.xlsx"

        # --- 4. Define default parameters for configurable stages ---
        self.stage1_params = {
            'spatialr': 25, 'ranger': 6, 'minsize': 200,
            'tilesizex': 4096, 'tilesizey': 4096, 'ram': 4096
        }
        self.stage2_params = {
            'learn_frac': 0.7, 'random_state': 42
        }
        self.stage4_params = {
            'classifier': 'rf',
            'rf_max': 110, 'rf_min': 2, 'rf_var': 16, 'rf_cat': 16, 'rf_acc': 0.01,
            'svm_c': 1.0, 'svm_k': 'linear'
        }

        # --- 5. Shared state variables ---
        self.feat_str = ""

    # --- Utility Methods ---

    def _ensure_directories(self):
        """Helper to enforce directory existence before writing files."""
        for d in [self.samples_dir, self.model_dir, self.seg_dir, self.class_dir]:
            d.mkdir(parents=True, exist_ok=True)

    def _run_cmd(self, cmd, stage, desc):
        """Executes a subprocess command with OTB environment variables set."""
        print(f"[Stage {stage}/{self.total_stages}] {desc}")

        # --- Configure OTB Environment Variables ---
        env = os.environ.copy()
        otb_bin = str(otb_dir / "bin")
        otb_lib = str(otb_dir / "lib")
        otb_apps = str(otb_dir / "lib" / "otb" / "applications")

        # Windows requires strictly separated paths
        # Prepend OTB paths to system PATH
        env["PATH"] = f"{otb_bin};{otb_lib};{env['PATH']}"

        env["OTB_APPLICATION_PATH"] = otb_apps
        env["GEOTIFF_CSV"] = str(otb_dir / "share" / "epsg_csv")
        env["GDAL_DATA"] = str(otb_dir / "share" / "gdal")

        # Run command
        proc = subprocess.Popen(cmd, shell=True, stdout=sys.stdout, stderr=sys.stderr, env=env)
        proc.communicate()
        if proc.returncode != 0:
            raise RuntimeError(f"Stage {stage} failed with return code {proc.returncode}: {cmd}")
        print(f"Completed stage {stage}/{self.total_stages}\n")

    def _resolve_raster(self, hdr):
        """Finds the actual image file associated with the header."""
        # 1. Check for standard extensions
        for ext in ['.img', '.tif', '.TIF']:
            p = hdr.with_suffix(ext)
            if p.exists(): return p

        # 2. Check for matching file with NO extension
        p_no_ext = hdr.with_suffix('')
        if p_no_ext.exists() and p_no_ext.is_file():
            return p_no_ext

        # 3. If the input WAS the image (e.g. .tif), return it
        if hdr.suffix in ['.tif', '.img']:
            return hdr

        raise FileNotFoundError(f"No raster image (.img/.tif) found matching header {hdr.stem}")

    def _raster_to_cutline(self, input_tif, cutline_shp, stage):
        print(f"[Stage {stage}/{self.total_stages}] Generating Cutline: Intersection of Raster Data & NUTS AOI")

        # --- Step 1: define NUTS path based on your structure ---
        # Structure: auxiliary_files/shapefiles_nuts/AT/NUTS2_AT.shp
        nuts_filename = f"NUTS2_{self.country}.shp"
        nuts_path = self.aux_dir / 'shapefiles_nuts' / self.country / nuts_filename

        if not nuts_path.exists():
            print(f"WARNING: NUTS shapefile not found at {nuts_path}")
            print("Cannot clip to country border. Proceeding with raster footprint only.")
            has_nuts = False
        else:
            print(f"Found NUTS shapefile: {nuts_path}")
            has_nuts = True

        # --- Step 2: Create Raster Footprint (Handling NaNs) ---
        ds = gdal.Open(str(input_tif))
        if not ds: raise RuntimeError(f"Could not open {input_tif}")

        band = ds.GetRasterBand(1)
        arr = band.ReadAsArray()
        nodata = band.GetNoDataValue()
        if nodata is None: nodata = 0

        # Robust valid pixel check (Handles Float/NaN)
        if np.issubdtype(arr.dtype, np.floating):
            is_valid = (arr != nodata) & (~np.isnan(arr))
        else:
            is_valid = (arr != nodata)

        mask_arr = is_valid.astype(np.uint8)

        # Create temporary memory raster
        drv_mem = gdal.GetDriverByName('MEM')
        mask_ds = drv_mem.Create('', ds.RasterXSize, ds.RasterYSize, 1, gdal.GDT_Byte)
        mask_ds.SetGeoTransform(ds.GetGeoTransform())
        mask_ds.SetProjection(ds.GetProjection())
        mask_band = mask_ds.GetRasterBand(1)
        mask_band.WriteArray(mask_arr)
        mask_band.SetNoDataValue(0)

        # Sieve filter to clean noise
        gdal.SieveFilter(mask_band, None, mask_band, 100, 8, [], None)

        # --- Step 3: Polygonize the Raster Footprint ---
        # We save this to a temporary file first
        temp_footprint = self.proc_dir / "temp_footprint.shp"

        shp_drv = ogr.GetDriverByName('ESRI Shapefile')
        if os.path.exists(str(temp_footprint)): shp_drv.DeleteDataSource(str(temp_footprint))

        out_ds = shp_drv.CreateDataSource(str(temp_footprint))
        srs = osr.SpatialReference()
        srs.ImportFromWkt(ds.GetProjection())
        layer = out_ds.CreateLayer('footprint', srs=srs, geom_type=ogr.wkbPolygon)
        layer.CreateField(ogr.FieldDefn('DN', ogr.OFTInteger))

        gdal.Polygonize(mask_band, None, layer, 0, [], callback=None)
        out_ds.Destroy()

        # --- Step 4: Intersect with NUTS (The Logic from Old Code) ---
        try:
            gdf_footprint = gpd.read_file(str(temp_footprint))
            # Keep only valid parts (DN=1)
            gdf_footprint = gdf_footprint[gdf_footprint['DN'] == 1]

            if has_nuts:
                gdf_nuts = gpd.read_file(str(nuts_path))

                # Ensure Coordinate Systems match
                if gdf_footprint.crs != gdf_nuts.crs:
                    print(f"Reprojecting NUTS to match Raster CRS...")
                    gdf_nuts = gdf_nuts.to_crs(gdf_footprint.crs)

                # --- THE CLIPPING OPERATION ---
                print("Clipping Raster Footprint to NUTS boundary...")
                # overlay(how='intersection') keeps only the overlapping area
                gdf_cutline = gpd.overlay(gdf_footprint, gdf_nuts, how='intersection')
            else:
                gdf_cutline = gdf_footprint

            if gdf_cutline.empty:
                raise RuntimeError("Intersection resulted in empty cutline! Check if NUTS overlaps with Raster.")

            # Dissolve to 1 feature
            gdf_final = gdf_cutline.dissolve()
            gdf_final['DN'] = 1

            # Save Final Cutline
            gdf_final.to_file(str(cutline_shp))
            print(f"Cutline saved to {cutline_shp}\n")

        except Exception as e:
            print(f"Error during cutline processing: {e}")
        finally:
            # Cleanup temp file
            if temp_footprint.exists():
                for ext in ['.shp', '.shx', '.dbf', '.prj', '.cpg']:
                    f = temp_footprint.with_suffix(ext)
                    if f.exists(): os.remove(f)

    def _clip_and_mask(self, input_tif, mask_tif, cutline_shp, out_tif, stage):
        print(f"[Stage {stage}/{self.total_stages}] Clipping to cutline and applying arable mask")
        ds_clip = gdal.Warp('', str(input_tif), format='MEM', cutlineDSName=str(cutline_shp), cropToCutline=True,
                            dstNodata=-9999)
        gt = ds_clip.GetGeoTransform();
        proj = ds_clip.GetProjection()
        cols, rows = ds_clip.RasterXSize, ds_clip.RasterYSize
        arr = ds_clip.GetRasterBand(1).ReadAsArray()
        nd = ds_clip.GetRasterBand(1).GetNoDataValue() if ds_clip.GetRasterBand(
            1).GetNoDataValue() is not None else -9999

        minx, maxy = gt[0], gt[3]
        maxx = minx + cols * gt[1];
        miny = maxy + rows * gt[5]
        mem_m = gdal.Warp('', str(mask_tif), format='MEM', width=cols, height=rows,
                          outputBounds=(minx, miny, maxx, miny + rows * abs(gt[5])),
                          dstSRS=proj, resampleAlg=gdal.GRA_NearestNeighbour,
                          srcNodata=0, dstNodata=0)
        mask_arr = mem_m.GetRasterBand(1).ReadAsArray()

        out_arr = np.where((arr != nd) & (mask_arr == 1), arr, nd).astype(np.float32)
        driver = gdal.GetDriverByName('GTiff')
        out_ds = driver.Create(str(out_tif), cols, rows, 1, gdal.GDT_Float32)
        out_ds.SetGeoTransform(gt);
        out_ds.SetProjection(proj)
        bnd = out_ds.GetRasterBand(1);
        bnd.WriteArray(out_arr);
        bnd.SetNoDataValue(nd)
        out_ds.FlushCache()
        print(f"Completed stage {stage}/{self.total_stages}\n")

    # --- Stage 1: Segmentation ---
    def stage_1_segmentation(self, **kwargs):
        self._ensure_directories()
        params = self.stage1_params.copy()
        params.update(kwargs)
        stage = 1

        if not self.seg_shp.exists():
            cmd = (
                f"otbcli_LargeScaleMeanShift -in {self.ras} -spatialr {params['spatialr']} "
                f"-ranger {params['ranger']} -minsize {params['minsize']} "
                f"-tilesizex {params['tilesizex']} -tilesizey {params['tilesizey']} "
                f"-mode vector -mode.vector.out {self.seg_shp} "
                f"-cleanup false -ram {params['ram']}"
            )
            self._run_cmd(cmd, stage, 'Image segmentation')
        else:
            print(f"[Stage {stage}/{self.total_stages}] Segmentation exists, skipping\n")

    # --- Stage 2: Sample Split ---
    def stage_2_split_samples(self, **kwargs):
        self._ensure_directories()
        params = self.stage2_params.copy()
        params.update(kwargs)
        stage = 2

        print(
            f"[Stage {stage}/{self.total_stages}] Splitting samples ({params['learn_frac'] * 100:.0f}/{(1 - params['learn_frac']) * 100:.0f})")
        if not self.sample_shp.exists():
            print(f"ERROR: Input sample file not found: {self.sample_shp}")
            return

        gdf = gpd.read_file(str(self.sample_shp))
        learn = gdf.sample(frac=params['learn_frac'], random_state=params['random_state'])
        control = gdf.drop(learn.index)

        learn.to_file(str(self.learn_shp))
        control.to_file(str(self.control_shp))
        print(f"Completed stage {stage}. Total {len(gdf)}, Learn {len(learn)}, Control {len(control)}\n")

    # --- Stage 3: Selection (FIXED FOR NEW GEOPANDAS) ---
    def stage_3_selection(self):
        self._ensure_directories()
        stage = 3
        if not self.sel_shp.exists():
            print(f"[Stage {stage}/{self.total_stages}] Running sample selection")
            if not self.learn_shp.exists():
                print("ERROR: Learn samples not found. Run Stage 2 first.")
                return
            if not self.seg_shp.exists():
                print("ERROR: Segmentation not found. Run Stage 1 first.")
                return

            pts = gpd.read_file(self.learn_shp)
            polys = gpd.read_file(self.seg_shp)

            # Ensure CRS matches
            if polys.crs != pts.crs:
                print("Warning: CRS mismatch. Re-projecting segmentation to sample CRS.")
                polys = polys.to_crs(pts.crs)

            # --- FIX: Use 'predicate' for modern GeoPandas ---
            # Your error "unexpected keyword argument 'op'" means you have a new version.
            try:
                sel = gpd.sjoin(polys, pts, how='inner', predicate='intersects')
            except TypeError:
                # Fallback for very old versions (just in case)
                sel = gpd.sjoin(polys, pts, how='inner', op='intersects')

            sel.to_file(self.sel_shp)
            print(f"[Stage {stage}/{self.total_stages}] Selected {len(sel)} features\n")
        else:
            print(f"[Stage {stage}/{self.total_stages}] Selection exists, skipping\n")

    # --- Stage 4: Train Classifier ---
    def stage_4_train_classifier(self, **kwargs):
        self._ensure_directories()

        force_retrain = kwargs.pop('force_retrain', False)
        params = self.stage4_params.copy()
        params.update(kwargs)
        stage = 4

        if not self.sel_shp.exists():
            print("ERROR: Selected samples file not found. Run Stage 3 first.")
            return

        df_sel = gpd.read_file(self.sel_shp)
        feats = [c for c in df_sel.columns if c.startswith('meanB')]
        if not feats:
            print("ERROR: No features starting with 'meanB' found in selected shapefile.")
            return
        self.feat_str = ' '.join(feats)

        clf_name = params['classifier']
        model_fn = self.model_dir / f"{self.country}_{self.track}_model.{clf_name}"
        confmat_fn = self.model_dir / f"{self.country}_{self.track}_train_confmat.{clf_name}.csv"

        if force_retrain or not model_fn.exists() or os.path.getsize(model_fn) == 0:
            if force_retrain and model_fn.exists():
                print(f"[Stage {stage}/{self.total_stages}] Parameters changed. Forcing overwrite.")
            elif model_fn.exists() and os.path.getsize(model_fn) == 0:
                print(f"[Stage {stage}/{self.total_stages}] Empty model, retrain\n")

            otb_clf_name = 'libsvm' if clf_name == 'svm' else clf_name
            clf_str = f"-classifier {otb_clf_name}"

            if clf_name == 'rf':
                clf_str += (
                    f" -classifier.rf.max {params['rf_max']} -classifier.rf.min {params['rf_min']} "
                    f" -classifier.rf.var {params['rf_var']} -classifier.rf.cat {params['rf_cat']} "
                    f" -classifier.rf.acc {params['rf_acc']}"
                )
            elif clf_name == 'svm':
                clf_str += f" -classifier.libsvm.c {params.get('svm_c', 1.0)}"
                clf_str += f" -classifier.libsvm.k {params.get('svm_k', 'linear')}"

            cmd = (
                f"otbcli_TrainVectorClassifier -io.vd {self.sel_shp} -io.out {model_fn} "
                f"-feat {self.feat_str} -cfield crop_id {clf_str} "
                f"-io.confmatout {confmat_fn}"
            )
            self._run_cmd(cmd, stage, f'Train {clf_name.upper()}')

            # --- Read and print the confusion matrix ---
            try:
                if confmat_fn.exists():
                    print(f"\n--- Training Confusion Matrix ---")
                    df_cm = pd.read_csv(confmat_fn, skiprows=2, index_col=0)
                    df_cm = df_cm.drop(index=['Total', 'UA'], columns=['Total', 'PA'], errors='ignore')
                    df_cm = df_cm.dropna(how='all', axis=0).dropna(how='all', axis=1)
                    df_cm = df_cm.fillna(0)
                    print(df_cm.to_string())
            except Exception as e:
                print(f"Warning: Could not read training confusion matrix: {e}\n")

        else:
            print(f"[Stage {stage}/{self.total_stages}] Model exists, skipping\n")

    # --- Stage 5: Classification ---
    def stage_5_classify_vector(self):
        self._ensure_directories()
        stage = 5

        clf_name = self.stage4_params['classifier']
        model_file = self.model_dir / f"{self.country}_{self.track}_model.{clf_name}"

        if not model_file.exists():
            print(f"ERROR: Model file {model_file} not found. Run Stage 4 first.")
            return

        if not self.feat_str:
            print("Feature string not set. Reading from selected samples...")
            try:
                if not self.sel_shp.exists():
                    raise FileNotFoundError("sel_shp not found, run stage 3")
                df_sel = gpd.read_file(self.sel_shp)
                feats = [c for c in df_sel.columns if c.startswith('meanB')]
                self.feat_str = ' '.join(feats)
            except Exception as e:
                print(f"ERROR: Could not determine features: {e}.")
                return

        if not self.class_shp.exists():
            cmd = (
                f"otbcli_VectorClassifier -in {self.seg_shp} -out {self.class_shp} "
                f"-model {model_file} -feat {self.feat_str} -cfield predicted -confmap true"
            )
            self._run_cmd(cmd, stage, 'Vector classification')
        else:
            print(f"[Stage {stage}/{self.total_stages}] Classification exists, skipping\n")

    # --- Stage 6: Rasterize Class ---
    def stage_6_rasterize_class(self):
        self._ensure_directories()
        stage = 6
        if not self.class_shp.exists():
            print(f"ERROR: Classified shapefile not found. Run Stage 5.")
            return
        if not self.class_tif.exists():
            cmd = (
                f"otbcli_Rasterization -in {self.class_shp} -out {self.class_tif} "
                f"-mode attribute -mode.attribute.field predicted -spx 10 -spy 10"
            )
            self._run_cmd(cmd, stage, 'Rasterize classification')
        else:
            print(f"[Stage {stage}/{self.total_stages}] Classified TIFF exists, skipping\n")

    # --- Stage 7: Rasterize Confidence ---
    def stage_7_rasterize_confidence(self):
        self._ensure_directories()
        stage = 7
        if not self.class_shp.exists():
            print(f"ERROR: Classified shapefile not found.")
            return
        if not self.conf_tif.exists():
            cmd = (
                f"otbcli_Rasterization -in {self.class_shp} -out {self.conf_tif} "
                f"-mode attribute -mode.attribute.field confidence -spx 10 -spy 10"
            )
            self._run_cmd(cmd, stage, 'Rasterize confidence')
        else:
            print(f"[Stage {stage}/{self.total_stages}] Confidence TIFF exists, skipping\n")

    # --- Stage 8: Cutline ---
    def stage_8_create_cutline(self):
        stage = 8
        if not self.cutline_shp.exists():
            self._raster_to_cutline(self.ras, self.cutline_shp, stage)
        else:
            print(f"[Stage {stage}/{self.total_stages}] Cutline exists, skipping\n")

    # --- Stage 9: Mask Class ---
    def stage_9_mask_class(self):
        self._ensure_directories()
        stage = 9
        mask_file = self.aux_dir / 'raster_files' / 'EU_arable_areas_mask_3857.tif'
        if not self.class_tif.exists():
            print(f"ERROR: Classified TIF not found.")
            return
        if not self.cutline_shp.exists():
            print(f"ERROR: Cutline not found.")
            return
        if not mask_file.exists():
            print(f"ERROR: Arable mask {mask_file} not found.")
            return

        if not self.masked_class.exists():
            self._clip_and_mask(self.class_tif, mask_file, self.cutline_shp, self.masked_class, stage)
        else:
            print(f"[Stage {stage}/{self.total_stages}] Masked classification exists, skipping\n")

    # --- Stage 10: Mask Confidence ---
    def stage_10_mask_confidence(self):
        self._ensure_directories()
        stage = 10
        mask_file = self.aux_dir / 'raster_files' / 'EU_arable_areas_mask_3857.tif'
        if not self.conf_tif.exists():
            print(f"ERROR: Confidence TIF not found.")
            return
        if not self.cutline_shp.exists():
            print(f"ERROR: Cutline not found.")
            return

        if not self.masked_conf.exists():
            self._clip_and_mask(self.conf_tif, mask_file, self.cutline_shp, self.masked_conf, stage)
        else:
            print(f"[Stage {stage}/{self.total_stages}] Masked confidence exists, skipping\n")

    # --- Stage 11: Metrics ---
    def stage_11_calculate_metrics(self):
        self._ensure_directories()
        stage = 11
        if not self.metrics_fp.exists():
            print(f"[Stage {stage}/{self.total_stages}] Computing metrics and exporting Excel")

            if not self.control_shp.exists():
                print(f"ERROR: Control shapefile not found.")
                return
            if not self.masked_class.exists():
                print(f"ERROR: Masked classification not found.")
                return

            # Read control points and masked classified raster
            ctrl = gpd.read_file(str(self.control_shp))
            ds = gdal.Open(str(self.masked_class))
            arr = ds.GetRasterBand(1).ReadAsArray()
            gt = ds.GetGeoTransform()
            inv = gdal.InvGeoTransform(gt)
            true_vals, pred_vals = [], []

            for _, row in ctrl.iterrows():
                try:
                    px = int(inv[0] + inv[1] * row.geometry.x + inv[2] * row.geometry.y)
                    py = int(inv[3] + inv[4] * row.geometry.x + inv[5] * row.geometry.y)
                    if 0 <= px < arr.shape[1] and 0 <= py < arr.shape[0]:
                        t = int(row['crop_id'])
                        p = int(arr[py, px])
                        if t > 0 and p > 0:
                            true_vals.append(t)
                            pred_vals.append(p)
                except Exception as e:
                    pass

            if not true_vals or not pred_vals:
                print("ERROR: No valid matching true/predicted values found.")
                return

            labels = sorted(list(set(true_vals + pred_vals)))
            cm = confusion_matrix(true_vals, pred_vals, labels=labels)
            precisions, recalls, f1s, _ = precision_recall_fscore_support(
                true_vals, pred_vals, labels=labels, average=None, zero_division=0
            )

            total = np.sum(cm)
            oa = np.trace(cm) / total
            sum_po = oa
            sum_pe = np.sum(np.sum(cm, axis=0) * np.sum(cm, axis=1)) / (total ** 2)
            kappa = (sum_po - sum_pe) / (1 - sum_pe) if (1 - sum_pe) != 0 else np.nan

            # compute area per class
            resx, resy = abs(gt[1]), abs(gt[5])
            area_ha = resx * resy / 10000
            unique_classes, counts = np.unique(arr[arr > 0], return_counts=True)
            class_areas = dict(zip(unique_classes, counts))
            areas = [{'Class': c, 'Area_ha': round(class_areas.get(c, 0) * area_ha, 2)} for c in labels]

            # Write Excel
            wb = openpyxl.Workbook()
            sh = wb.active
            sh.title = 'Results'

            # Confusion Matrix
            sh.cell(row=1, column=1, value='Confusion Matrix').font = Font(bold=True)
            sh.cell(row=2, column=1, value='True \\ Pred').font = Font(bold=True)
            for j, lbl in enumerate(labels, start=2):
                sh.cell(row=2, column=j, value=lbl).font = Font(bold=True)
            for i, lbl in enumerate(labels, start=3):
                sh.cell(row=i, column=1, value=lbl).font = Font(bold=True)
                for j, _ in enumerate(labels):
                    sh.cell(row=i, column=j + 2, value=int(cm[i - 3, j]))

            # OA & Kappa
            base = 4 + len(labels)
            sh.cell(row=base, column=1, value='Overall Accuracy').font = Font(bold=True)
            sh.cell(row=base, column=2, value=round(oa, 4))
            sh.cell(row=base + 1, column=1, value='Kappa').font = Font(bold=True)
            sh.cell(row=base + 1, column=2, value=round(kappa, 4))

            # Classification metrics
            start = base + 3
            headers = ['Class', 'Producer Acc (Recall)', 'User Acc (Precision)', 'F1-score']
            for j, h in enumerate(headers, start=1):
                sh.cell(row=start, column=j, value=h).font = Font(bold=True)
            for idx, c in enumerate(labels):
                row_idx = start + 1 + idx
                sh.cell(row=row_idx, column=1, value=c)
                sh.cell(row=row_idx, column=2, value=round(recalls[idx], 4))
                sh.cell(row=row_idx, column=3, value=round(precisions[idx], 4))
                sh.cell(row=row_idx, column=4, value=round(f1s[idx], 4))

            # Areas
            ar0 = start + 1 + len(labels) + 1
            sh.cell(row=ar0, column=1, value='Areas (ha)').font = Font(bold=True)
            sh.cell(row=ar0 + 1, column=1, value='Class').font = Font(bold=True)
            sh.cell(row=ar0 + 1, column=2, value='Area_ha').font = Font(bold=True)
            for idx, a in enumerate(areas, start=ar0 + 2):
                sh.cell(row=idx, column=1, value=a['Class'])
                sh.cell(row=idx, column=2, value=a['Area_ha'])

            wb.save(str(self.metrics_fp))
            print(f"Metrics and Excel saved to {self.metrics_fp}\n")
        else:
            print(f"[Stage 11/{self.total_stages}] Metrics Excel exists, skipping")
        print(f"All done! Metrics available at {self.metrics_fp}")


# --- Interactive Menu Helpers ---

def get_params(param_dict):
    """Helper to interactively update a parameter dictionary."""
    new_params = param_dict.copy()
    print("--- Current Parameters ---")
    for key, val in new_params.items():
        print(f"  {key}: {val}")

    if input("Change parameters? (y/n) [n]: ").lower() != 'y':
        return new_params

    for key, val in new_params.items():
        new_val_str = input(f"Enter new value for '{key}' [{val}]: ")
        if not new_val_str:
            continue

        try:
            original_type = type(val)
            new_params[key] = original_type(new_val_str)
        except ValueError:
            print(f"Invalid value. Keeping default {val}.")

    print("--- Updated Parameters ---")
    for key, val in new_params.items():
        print(f"  {key}: {val}")
    return new_params


def get_classifier_params(param_dict):
    """Special helper for classifier params."""
    new_params = param_dict.copy()
    print("--- Current Parameters ---")
    for key, val in new_params.items():
        print(f"  {key}: {val}")

    if input("Change parameters? (y/n) [n]: ").lower() != 'y':
        return new_params

    clf = input(f"Enter classifier (e.g., rf, svm) [{new_params['classifier']}]: ") or new_params['classifier']
    new_params['classifier'] = clf.lower()

    print(f"\n--- Setting parameters for {clf.upper()} ---")
    prefix = 'rf_' if clf == 'rf' else 'svm_' if clf == 'svm' else None

    if prefix:
        for key in [k for k in new_params if k.startswith(prefix)]:
            val = new_params[key]
            if key == 'svm_k':
                options = ['linear', 'rbf', 'poly', 'sigmoid']
                new_val_str = input(f"Enter new value for '{key}' ({'/'.join(options)}) [{val}]: ")
                if not new_val_str: new_val_str = str(val)
                if new_val_str in options:
                    new_params[key] = new_val_str
            else:
                new_val_str = input(f"Enter new value for '{key}' [{val}]: ")
                if new_val_str:
                    try:
                        new_params[key] = type(val)(new_val_str)
                    except ValueError:
                        print(f"Invalid value.")

    return new_params


# --- Main Execution ---

def main_menu(pipeline):
    """Displays the main interactive menu."""
    menu = f"""
    --- Processing Pipeline Menu ---
    Track: {pipeline.track} ({pipeline.country})

    [1] Stage 1: Segmentation
    [2] Stage 2: Split Samples (current: {pipeline.stage2_params['learn_frac'] * 100:.0f}%)
    [3] Stage 3: Select Training Polygons
    [4] Stage 4: Train Classifier (current: {pipeline.stage4_params['classifier'].upper()})
    [5] Stage 5: Classify Vector
    [6] Stage 6: Rasterize Classification
    [7] Stage 7: Rasterize Confidence
    [8] Stage 8: Create Valid-Pixel Cutline
    [9] Stage 9: Mask Classification
    [10] Stage 10: Mask Confidence
    [11] Stage 11: Calculate Metrics

    [A] Run All Stages (using current parameters)
    [Q] Quit

    Enter your choice:
    """

    while True:
        choice = input(menu).strip().upper()
        try:
            if choice == '1':
                new_params = get_params(pipeline.stage1_params)
                pipeline.stage1_params.update(new_params)
                pipeline.stage_1_segmentation(**pipeline.stage1_params)
            elif choice == '2':
                new_params = get_params(pipeline.stage2_params)
                pipeline.stage2_params.update(new_params)
                pipeline.stage_2_split_samples(**pipeline.stage2_params)
            elif choice == '3':
                pipeline.stage_3_selection()
            elif choice == '4':
                new_params = get_classifier_params(pipeline.stage4_params)
                force = (pipeline.stage4_params != new_params)
                pipeline.stage4_params.update(new_params)
                pipeline.stage_4_train_classifier(force_retrain=force, **pipeline.stage4_params)
            elif choice == '5':
                pipeline.stage_5_classify_vector()
            elif choice == '6':
                pipeline.stage_6_rasterize_class()
            elif choice == '7':
                pipeline.stage_7_rasterize_confidence()
            elif choice == '8':
                pipeline.stage_8_create_cutline()
            elif choice == '9':
                pipeline.stage_9_mask_class()
            elif choice == '10':
                pipeline.stage_10_mask_confidence()
            elif choice == '11':
                pipeline.stage_11_calculate_metrics()
            elif choice == 'A':
                pipeline.stage_1_segmentation(**pipeline.stage1_params)
                pipeline.stage_2_split_samples(**pipeline.stage2_params)
                pipeline.stage_3_selection()
                pipeline.stage_4_train_classifier(**pipeline.stage4_params)
                pipeline.stage_5_classify_vector()
                pipeline.stage_6_rasterize_class()
                pipeline.stage_7_rasterize_confidence()
                pipeline.stage_8_create_cutline()
                pipeline.stage_9_mask_class()
                pipeline.stage_10_mask_confidence()
                pipeline.stage_11_calculate_metrics()
            elif choice == 'Q':
                break
        except Exception as e:
            print(f"\n--- ERROR ---: {e}")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Modular OBIA Classification Pipeline")
    parser.add_argument('--track', required=True, help="Processing track ID (e.g., P1, P2)")
    args = parser.parse_args()

    try:
        pipeline = ProcessingPipeline(track=args.track)
        main_menu(pipeline)
    except Exception as e:
        print(f"Initialization Error: {e}")
        sys.exit(1)