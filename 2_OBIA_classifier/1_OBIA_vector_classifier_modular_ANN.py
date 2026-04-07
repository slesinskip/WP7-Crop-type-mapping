import os
import argparse
from pathlib import Path
import subprocess
import sys
import shlex
import geopandas as gpd
import numpy as np
import pandas as pd
from osgeo import gdal, ogr, osr, gdalconst
from sklearn.metrics import confusion_matrix, precision_recall_fscore_support
from sklearn.neural_network import MLPClassifier
from sklearn.preprocessing import StandardScaler
from sklearn.utils import resample
import joblib
import openpyxl
from openpyxl.styles import Font

from concurrent.futures import ThreadPoolExecutor
import threading

# Try importing scikit-image
try:
    from skimage.segmentation import felzenszwalb, slic
    from skimage.util import img_as_float
    from skimage.measure import regionprops_table

    HAS_SKIMAGE = True
except ImportError:
    HAS_SKIMAGE = False
    print("WARNING: scikit-image not found. This script requires scikit-image for raster segmentation.")

# --- Configuration (Global) ---

# python 1_OBIA_vector_classifier_modular_ANN.py --track P1a

# Base Paths provided by user
base_dir = Path("D:/AIML_CropMapper_Cloud/workingDir")
aux_dir = Path("D:/AIML_CropMapper_Cloud/auxiliary_files")

# OTB Installation Path (Still used for some auxiliary tasks if needed, but main flow is Python)
otb_dir = Path("D:/AIML_CropMapper_Cloud/2_OBIA_classifier/OTB-6.2.0-Win64")

# Track to Country Mapping
track_regions = {
    'P1': 'AT', 'P1a': 'AT',
    'P2': 'IE', 'P2a': 'IE',
    'P3': 'NL',
    'P4': 'PT', 'P4a': 'PT'
}
TOTAL_STAGES = 8


# --- Main Pipeline Class ---

class ProcessingPipeline:
    def __init__(self, track):
        self.track = track
        try:
            self.country = track_regions[track]
        except KeyError:
            print(f"Error: Track '{track}' not defined in track_regions configuration.")
            sys.exit(1)

        self.total_stages = TOTAL_STAGES
        print(f"Initializing pipeline for Track: {self.track}, Country: {self.country}")

        # --- 1. Define all paths ---
        self.base_dir = base_dir
        self.aux_dir = aux_dir
        self.proc_dir = self.base_dir / self.track / 'processed_raster'
        self.out_dir = self.base_dir / self.track / 'classification_results'
        self.samples_dir = self.out_dir / 'samples'
        self.model_dir = self.out_dir / 'train_model'
        self.seg_dir = self.out_dir / 'segmentation'
        self.class_dir = self.out_dir / 'classification'

        self._ensure_directories()

        # --- 2. Resolve input raster ---
        search_patterns = [
            f"{self.track}_*_VH_VV*.tif",
            f"*_{self.track}_*_VH_VV*.tif",
            f"*{self.track}*.tif",
            f"{self.track}_*_VH_VV*.hdr",
            f"*{self.track}*.hdr",
        ]

        self.hdr = None
        if self.proc_dir.exists():
            for pattern in search_patterns:
                self.hdr = next(self.proc_dir.glob(pattern), None)
                if self.hdr:
                    break

            if not self.hdr:
                raise FileNotFoundError(f"No raster file (TIF/HDR) found for track {self.track} in {self.proc_dir}")

            self.ras = self._resolve_raster(self.hdr)
            print(f"Input raster found: {self.ras}")
        else:
            raise FileNotFoundError(f"Processing directory does not exist: {self.proc_dir}")

        # --- 3. Define all output file paths ---
        # Zmieniamy rozszerzenie wektora na .sqlite, aby ominąć 2GB limit dla Shapefile!
        self.seg_tif = self.seg_dir / f"{self.country}_{self.track}_segmentation.tif"
        self.seg_shp = self.seg_dir / f"{self.country}_{self.track}_segmentation.sqlite"

        # Samples
        samples_base = self.aux_dir / 'shapefiles_samples'
        candidate_paths = [
            samples_base / f"{self.country}_{self.track}" / "samples.shp",
            samples_base / self.track / "samples.shp",
            samples_base / self.country / "samples.shp",
            samples_base / "samples.shp"
        ]

        self.sample_shp = None
        for p in candidate_paths:
            if p.exists():
                self.sample_shp = p
                print(f"Training samples found at: {self.sample_shp}")
                break

        if not self.sample_shp:
            print(f"\nCRITICAL WARNING: Could not find 'samples.shp' inside {samples_base}")
            self.sample_shp = samples_base / f"{self.country}_{self.track}" / "samples.shp"

        # Output paths
        self.learn_shp = self.samples_dir / 'learn.shp'
        self.control_shp = self.samples_dir / 'control.shp'
        self.sel_csv = self.samples_dir / f"{self.country}_{self.track}_learn_features.csv"

        # Classification outputs
        self.class_tif = self.class_dir / f"{self.country}_{self.track}_classified.tif"
        self.conf_tif = self.class_dir / f"{self.country}_{self.track}_confidence_map.tif"

        self.masked_class = self.class_dir / f"{self.country}_{self.track}_classified_masked.tif"
        self.masked_conf = self.class_dir / f"{self.country}_{self.track}_confidence_masked.tif"
        self.metrics_fp = self.class_dir / f"{self.country}_{self.track}_metrics.xlsx"

        # --- 4. Parameters ---
        self.stage1_params = {
            'method': 'otb_meanshift',
            'tile_size': 4096,
            'ram': 65536,

            # OTB Params (User's original working params for LargeScaleMeanShift vector mode)
            # Zastosowano ranger 2.0, optymalny dla stert dB radaru, oraz 1024 tiles do unikania obcietych map
            'spatialr': 20, 'ranger': 6.0, 'minsize': 100, 'tilesizex': 4096, 'tilesizey': 4096,

            # Python Params (Fallback)
            'n_segments': 20000, 'compactness': 5.0, 'slic_sigma': 1.0,
            'scale': 100, 'sigma': 0.8, 'min_size': 50
        }
        self.stage2_params = {
            'learn_frac': 0.7, 'random_state': 42
        }
        self.stage4_params = {
            'classifier': 'ann_sklearn',
            'sk_hidden_sizes': '100,50',
            'sk_activation': 'relu',
            'sk_solver': 'adam',
            'sk_alpha': 0.0001,
            'sk_max_iter': 500,
            'balance_threshold': 1000
        }

        self.feat_cols = []

        # --- Utility Methods ---

    def _ensure_directories(self):
        for d in [self.samples_dir, self.model_dir, self.seg_dir, self.class_dir]:
            d.mkdir(parents=True, exist_ok=True)

    def _run_cmd(self, cmd, stage, desc, ram=None):
        print(f"[Stage {stage}/{self.total_stages}] {desc}")
        env = os.environ.copy()
        otb_bin = str(otb_dir / "bin")
        otb_lib = str(otb_dir / "lib")
        otb_apps = str(otb_dir / "lib" / "otb" / "applications")
        env["PATH"] = f"{otb_bin};{otb_lib};{env['PATH']}"
        env["OTB_APPLICATION_PATH"] = otb_apps
        env["GEOTIFF_CSV"] = str(otb_dir / "share" / "epsg_csv")
        env["GDAL_DATA"] = str(otb_dir / "share" / "gdal")

        if ram:
            env["OTB_MAX_RAM_HINT"] = str(ram)

        if isinstance(cmd, str):
            cmd = shlex.split(cmd, posix=os.name != 'nt')
        proc = subprocess.Popen(cmd, shell=False, stdout=sys.stdout, stderr=sys.stderr, env=env)
        proc.communicate()
        if proc.returncode != 0:
            raise RuntimeError(f"Stage {stage} failed with return code {proc.returncode}: {cmd}")
        print(f"Completed stage {stage}/{self.total_stages}\n")

    def _resolve_raster(self, hdr):
        if hdr.suffix.lower() in ['.tif', '.tiff']: return hdr
        for ext in ['.img', '.tif', '.TIF']:
            p = hdr.with_suffix(ext)
            if p.exists(): return p
        p_no_ext = hdr.with_suffix('')
        if p_no_ext.exists() and p_no_ext.is_file(): return p_no_ext
        raise FileNotFoundError(f"No raster image (.img/.tif) found matching header {hdr.stem}")

    def _apply_mask(self, input_tif, mask_tif, out_tif, stage):
        print(f"[Stage {stage}/{self.total_stages}] Applying Arable & Data Footprint Mask...")

        # Open the original radar stack to check for data footprint (blank areas)
        ds_stack = gdal.Open(str(self.ras))
        if not ds_stack: raise RuntimeError(f"Could not open source raster {self.ras} for footprint masking.")
        stack_band = ds_stack.GetRasterBand(1)

        if not mask_tif.exists():
            print(f"    WARNING: Arable mask not found at {mask_tif}. Will only apply data footprint mask.")
            has_arable_mask = False
        else:
            has_arable_mask = True

        ds_in = gdal.Open(str(input_tif))
        gt = ds_in.GetGeoTransform()
        proj = ds_in.GetProjection()
        cols = ds_in.RasterXSize
        rows = ds_in.RasterYSize

        minx = gt[0]
        maxy = gt[3]
        maxx = minx + gt[1] * cols
        miny = maxy + gt[5] * rows

        if has_arable_mask:
            # Warp the mask to match the input raster exactly
            temp_mask_vrt = str(out_tif).replace('.tif', '_mask_temp.vrt')
            mask_opts = gdal.WarpOptions(
                format='VRT',
                outputBounds=(minx, miny, maxx, maxy),
                width=cols,
                height=rows,
                dstSRS=proj,
                resampleAlg=gdal.GRA_NearestNeighbour
            )
            ds_mask = gdal.Warp(temp_mask_vrt, str(mask_tif), options=mask_opts)
            if not ds_mask: raise RuntimeError("Failed to warp the arable mask.")
            m_band = ds_mask.GetRasterBand(1)
        else:
            ds_mask = None

        in_band = ds_in.GetRasterBand(1)
        out_type = in_band.DataType

        driver = gdal.GetDriverByName('GTiff')
        out_ds = driver.Create(str(out_tif), cols, rows, 1, out_type,
                               options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
        out_ds.SetGeoTransform(gt)
        out_ds.SetProjection(proj)
        out_band = out_ds.GetRasterBand(1)

        nodata = in_band.GetNoDataValue()
        if nodata is None: nodata = 0
        out_band.SetNoDataValue(nodata)

        tile_size = 4096

        for y in range(0, rows, tile_size):
            for x in range(0, cols, tile_size):
                xsize = min(tile_size, cols - x)
                ysize = min(tile_size, rows - y)

                # 1. Read the Classification/Confidence Data
                arr = in_band.ReadAsArray(x, y, xsize, ysize)

                # 2. Read the Original Radar Stack (to find empty track halves)
                # Ensure we handle potential out of bounds or missing stack bands gracefully
                if stack_band:
                    try:
                        stack_arr = stack_band.ReadAsArray(x, y, xsize, ysize)
                        # Apply Footprint Mask: If original radar data is exactly 0 or NaN, force output to NoData
                        if stack_arr is not None:
                            stack_nodata = stack_band.GetNoDataValue()
                            if stack_nodata is not None:
                                arr[stack_arr == stack_nodata] = nodata
                            arr[stack_arr == 0] = nodata
                            arr[np.isnan(stack_arr)] = nodata
                    except Exception as e:
                        print(f"Warning: Failed to read stack band: {e}")

                # 3. Read and Apply the Arable Mask (if it exists)
                if ds_mask:
                    m_arr = m_band.ReadAsArray(x, y, xsize, ysize)
                    arr[m_arr < 0.5] = nodata

                out_band.WriteArray(arr, x, y)

        out_ds.FlushCache()

        ds_mask = None
        out_ds = None
        ds_in = None
        ds_stack = None
        if has_arable_mask and os.path.exists(temp_mask_vrt): os.remove(temp_mask_vrt)

        print(f"Completed stage {stage}\n")

    def _create_seasonal_composite(self):
        """Creates a lightweight 6-band composite (3 evenly spaced dates x 2 pol) from the optimal 'Golden Window' of vegetation."""
        import re
        from datetime import datetime

        print("    [INFO] Analyzing radar stack to select the best 3 dates for the 'Golden Vegetation Window'...")

        # Define the Golden Window of vegetation contrast (May 15th to August 15th)
        start_month, start_day = 5, 15
        end_month, end_day = 8, 15

        # Open full raster to extract band dates
        ds = gdal.Open(str(self.ras))
        if not ds:
            raise RuntimeError(f"Could not open source raster {self.ras}")

        nbands = ds.RasterCount
        dates_bands = {}

        # Parse band names (e.g., "Sigma0_VH_05May2024")
        for i in range(1, nbands + 1):
            band = ds.GetRasterBand(i)
            desc = band.GetDescription()
            m = re.search(r"(\d{2}[A-Za-z]{3}\d{4})", desc)
            if m:
                date_str = m.group(1)
                try:
                    dt = datetime.strptime(date_str, "%d%b%Y")
                    if dt not in dates_bands:
                        dates_bands[dt] = []
                    dates_bands[dt].append(i)
                except ValueError:
                    pass

        if len(dates_bands) < 3:
            print("    [WARNING] Less than 3 dates found in the stack! Using full raster for segmentation.")
            return self.ras

        # Sort all discovered dates chronologically
        all_dates = sorted(list(dates_bands.keys()))

        # Find dates that fall strictly within the Golden Window (irrespective of year)
        golden_dates = []
        for d in all_dates:
            # Create comparable numeric representation of MM.DD
            day_val = d.month + d.day / 100.0
            start_val = start_month + start_day / 100.0
            end_val = end_month + end_day / 100.0

            if start_val <= day_val <= end_val:
                golden_dates.append(d)

        # Smart selection of exactly 3 dates
        selected_dates = []
        if len(golden_dates) >= 3:
            # We have plenty of optimal summer dates! Let's pick 3 evenly spaced dates
            # to capture maximum phenomenological difference (start, middle, end of window).
            idx_step = (len(golden_dates) - 1) / 2.0
            selected_dates = [golden_dates[int(round(0))],
                              golden_dates[int(round(idx_step))],
                              golden_dates[-1]]
            print(f"    [INFO] Found {len(golden_dates)} dates in the Golden Window (Mid-May to Mid-Aug).")
        elif len(golden_dates) > 0:
            # We have 1 or 2 dates in the summer, but we need 3 total.
            # Pad the rest with the latest available dates just before the window.
            print(
                f"    [INFO] Found only {len(golden_dates)} dates in the Golden Window. Padding with latest spring dates.")
            needed = 3 - len(golden_dates)
            prior_dates = [d for d in all_dates if d not in golden_dates and d < golden_dates[0]]
            if len(prior_dates) >= needed:
                selected_dates = prior_dates[-needed:] + golden_dates
            else:
                # If everything fails, just take the 3 most recent dates overall
                selected_dates = all_dates[-3:]
        else:
            # Time series doesn't reach May 15th at all (e.g. early mapping ending in April).
            # Just take the 3 most recent dates available as they represent the most mature growth.
            print("    [INFO] No dates found in the Summer Golden Window. Selecting the 3 most recent dates available.")
            selected_dates = all_dates[-3:]

        # Sort chronologically just to be safe
        selected_dates.sort()

        selected_bands = []
        for d in selected_dates:
            selected_bands.extend(dates_bands[d])

        print(
            f"    [INFO] Final 3 dates chosen for segmentation composite: {[d.strftime('%Y-%m-%d') for d in selected_dates]}")
        print(f"    [INFO] Extracting {len(selected_bands)} bands (VH+VV) into a lightweight composite...")

        composite_tif = self.seg_dir / f"{self.country}_{self.track}_seasonal_composite.tif"

        if not composite_tif.exists():
            gdal.Translate(
                str(composite_tif),
                str(self.ras),
                bandList=selected_bands,
                format='GTiff',
                creationOptions=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES']
            )
            print(f"    [INFO] Seasonal composite saved to {composite_tif}")
        else:
            print(f"    [INFO] Seasonal composite already exists.")

        return composite_tif

    # --- Stage 1: OTB Segmentation (Direct Raster Pipeline) ---
    def stage_1_segmentation(self, **kwargs):
        self._ensure_directories()
        params = self.stage1_params.copy()
        params.update(kwargs)
        stage = 1

        if self.seg_tif.exists():
            print(f"[Stage {stage}/{self.total_stages}] Segmentation Raster exists, skipping\n")
            return

        method = params.get('method', 'otb_meanshift')

        if method in ['otb_meanshift', 'otb_meanshift_seasonal']:
            print(f"[Stage {stage}/{self.total_stages}] Running OTB Large-Scale Mean-Shift (Vector Mode) [{method}]...")

            input_raster_for_seg = self.ras
            if method == 'otb_meanshift_seasonal':
                try:
                    input_raster_for_seg = self._create_seasonal_composite()
                except Exception as e:
                    print(f"    [WARNING] Failed to create seasonal composite: {e}. Falling back to full stack.")
                    input_raster_for_seg = self.ras

            if not self.seg_shp.exists():
                cmd = (
                    f"otbcli_LargeScaleMeanShift -in {input_raster_for_seg} -spatialr {params['spatialr']} "
                    f"-ranger {params['ranger']} -minsize {params['minsize']} "
                    f"-tilesizex {params['tilesizex']} -tilesizey {params['tilesizey']} "
                    f"-mode vector -mode.vector.out {self.seg_shp} "
                    f"-cleanup false -ram {params['ram']}"
                )
                self._run_cmd(cmd, stage, 'OTB LargeScaleMeanShift (Vector)')
                if not self.seg_shp.exists():
                    raise RuntimeError("OTB Vector Segmentation failed: output shapefile not found.")
            else:
                print(f"    [INFO] Segmentation vector already exists at {self.seg_shp}")

            # --- Rasterize the vector shapefile so ANN processing can use it as a tiled grid ---
            if not self.seg_tif.exists():
                print(f"    [INFO] Rasterizing segmentation for ANN feature extraction...")

                ds_stack = gdal.Open(str(input_raster_for_seg))
                gt = ds_stack.GetGeoTransform()
                proj = ds_stack.GetProjection()
                cols = ds_stack.RasterXSize
                rows = ds_stack.RasterYSize

                # OTB generates a column 'DN' or 'label' for region IDs. Typically 'DN' for vector mode in old versions, or 'label'.
                # gdal.Rasterize will burn the object ID into the pixels.

                driver = gdal.GetDriverByName('GTiff')
                out_ds = driver.Create(str(self.seg_tif), cols, rows, 1, gdal.GDT_Int32,
                                       options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
                out_ds.SetGeoTransform(gt)
                out_ds.SetProjection(proj)

                # Check what field OTB generated
                shp_ds = ogr.Open(str(self.seg_shp))
                layer = shp_ds.GetLayer()
                field_names = [field.name for field in layer.schema]
                id_field = 'DN' if 'DN' in field_names else 'label' if 'label' in field_names else field_names[0]

                print(f"    [INFO] Burning field '{id_field}' into raster...")
                gdal.RasterizeLayer(out_ds, [1], layer, options=[f"ATTRIBUTE={id_field}"])

                out_ds.FlushCache()
                out_ds = None
                shp_ds = None
                ds_stack = None
                print(f"    [INFO] Rasterization complete: {self.seg_tif}")
            else:
                print(f"    [INFO] Rasterized segmentation already exists.")

            return

        if method in ['python_felzenszwalb', 'python_slic']:
            if not HAS_SKIMAGE:
                print("Error: scikit-image not installed.")
                return
            self._run_python_segmentation_tiled(params, stage, method)
            return

        print(f"Error: Unknown segmentation method '{method}'")

    def _run_python_segmentation_tiled(self, params, stage, method):
        print(f"[Stage {stage}/{self.total_stages}] Running Tiled Python Segmentation ({method})...")

        try:
            ds = gdal.Open(str(self.ras))
            if not ds: raise RuntimeError("Could not open raster")

            cols = ds.RasterXSize
            rows = ds.RasterYSize
            nbands = ds.RasterCount
            gt = ds.GetGeoTransform()
            proj = ds.GetProjection()

            driver = gdal.GetDriverByName('GTiff')
            out_ds = driver.Create(str(self.seg_tif), cols, rows, 1, gdal.GDT_Int32,
                                   options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
            out_ds.SetGeoTransform(gt)
            out_ds.SetProjection(proj)
            out_band = out_ds.GetRasterBand(1)
            out_band.SetNoDataValue(0)

            tile_size = params.get('tile_size', 4096)
            global_seg_id = 1

            for y in range(0, rows, tile_size):
                for x in range(0, cols, tile_size):
                    xsize = min(tile_size, cols - x)
                    ysize = min(tile_size, rows - y)

                    print(f"    Processing Tile: x={x}, y={y}")

                    img_list = []
                    for b in range(1, nbands + 1):
                        band = ds.GetRasterBand(b)
                        arr = band.ReadAsArray(x, y, xsize, ysize)
                        arr = np.nan_to_num(arr)
                        img_list.append(arr)

                    img = np.dstack(img_list)
                    if np.all(img == 0): continue

                    valid_mask = np.sum(np.abs(img), axis=2) > 0
                    if not np.any(valid_mask): continue

                    img_norm = img_as_float(img)

                    if method == 'python_felzenszwalb':
                        segments = felzenszwalb(img_norm, scale=params['scale'], sigma=params['sigma'],
                                                min_size=params['min_size'])
                    elif method == 'python_slic':
                        segments = slic(img_norm, n_segments=params['n_segments'], compactness=params['compactness'],
                                        sigma=params['slic_sigma'], start_label=1, mask=valid_mask)

                    seg_valid_mask = segments > 0
                    segments[seg_valid_mask] += global_seg_id
                    segments[~valid_mask] = 0

                    if np.any(seg_valid_mask):
                        global_seg_id = segments.max() + 1

                    out_band.WriteArray(segments.astype(np.int32), x, y)

            out_ds.FlushCache()
            out_ds = None
            print(f"    Segmentation Raster saved to {self.seg_tif}\n")

        except Exception as e:
            print(f"ERROR in Python segmentation: {e}")
            raise

    # --- Stage 2: Sample Split ---
    def stage_2_split_samples(self, **kwargs):
        self._ensure_directories()
        params = self.stage2_params.copy()
        params.update(kwargs)
        stage = 2

        if not self.sample_shp.exists():
            print("ERROR: Input sample file not found.")
            return

        gdf = gpd.read_file(str(self.sample_shp), engine="pyogrio")
        learn = gdf.sample(frac=params['learn_frac'], random_state=params['random_state'])
        control = gdf.drop(learn.index)

        learn.to_file(str(self.learn_shp), engine="pyogrio")
        control.to_file(str(self.control_shp), engine="pyogrio")
        print(f"Completed stage {stage}.\n")

    # --- Stage 3: Feature Extraction (Object-Based) ---
    def stage_3_selection(self):
        self._ensure_directories()
        stage = 3

        if self.sel_csv.exists():
            print(f"[Stage {stage}] Features already extracted, skipping.")
            return

        print(f"[Stage {stage}/{self.total_stages}] Extracting OBJECT-BASED features for Training Points...")

        if not self.learn_shp.exists():
            print("ERROR: Learn samples not found.")
            return

        gdf = gpd.read_file(str(self.learn_shp), engine="pyogrio")

        ds = gdal.Open(str(self.ras))
        gt = ds.GetGeoTransform()
        inv_gt = gdal.InvGeoTransform(gt)
        raster_proj = ds.GetProjection()
        nbands = ds.RasterCount
        cols = ds.RasterXSize
        rows = ds.RasterYSize

        # --- FIX: ALIGN CRS FOR TRAINING SAMPLES ---
        from pyproj import CRS
        if raster_proj and gdf.crs:
            target_crs = CRS.from_wkt(raster_proj)
            if gdf.crs != target_crs:
                print(f"    Warning: Reprojecting samples from {gdf.crs.name} to Match Raster CRS...")
                gdf = gdf.to_crs(target_crs)

        print(f"    Finding target segments for {len(gdf)} points...")

        seg_ds = gdal.Open(str(self.seg_tif))
        seg_band = seg_ds.GetRasterBand(1)

        target_segments = {}

        xs = gdf.geometry.x.values
        ys = gdf.geometry.y.values

        pxs = (inv_gt[0] + inv_gt[1] * xs + inv_gt[2] * ys).astype(int)
        pys = (inv_gt[3] + inv_gt[4] * xs + inv_gt[5] * ys).astype(int)
        crop_ids = gdf['crop_id'].values

        for px, py, crop_id in zip(pxs, pys, crop_ids):
            if 0 <= px < cols and 0 <= py < rows:
                try:
                    seg_id = seg_band.ReadAsArray(int(px), int(py), 1, 1)[0, 0]
                    if seg_id > 0:
                        target_segments[seg_id] = crop_id
                except:
                    pass

        if not target_segments:
            print("ERROR: No valid samples found overlapping the raster.")
            return

        print(f"    Found {len(target_segments)} unique segments for training.")
        print("    Calculating true segment means (Optimized Tiled Read)...")

        target_ids_set = set(target_segments.keys())

        sums = {tid: np.zeros(nbands, dtype=np.float64) for tid in target_ids_set}
        sums_sq = {tid: np.zeros(nbands, dtype=np.float64) for tid in target_ids_set}
        counts = {tid: 0 for tid in target_ids_set}

        tile_size = 2048

        for y in range(0, rows, tile_size):
            for x in range(0, cols, tile_size):
                xsize = min(tile_size, cols - x)
                ysize = min(tile_size, rows - y)

                seg_arr = seg_band.ReadAsArray(x, y, xsize, ysize)
                tile_ids = np.unique(seg_arr)

                intersect_ids = target_ids_set.intersection(tile_ids)

                if not intersect_ids:
                    continue

                sys.stdout.write(f"\r      Reading required data from Tile: x={x}, y={y}    ")
                sys.stdout.flush()

                stack_tile = ds.ReadAsArray(x, y, xsize, ysize)
                stack_tile = np.nan_to_num(stack_tile, copy=False)

                for tid in intersect_ids:
                    mask = (seg_arr == tid)
                    pixel_count = np.sum(mask)
                    counts[tid] += pixel_count
                    for b in range(nbands):
                        vals = stack_tile[b][mask]
                        sums[tid][b] += np.sum(vals)
                        sums_sq[tid][b] += np.sum(vals ** 2)

        print("\n    Aggregation complete. Formatting features...")

        valid_tids = [tid for tid in target_ids_set if counts[tid] > 0]

        crop_ids = [target_segments[tid] for tid in valid_tids]
        seg_ids = valid_tids

        if len(valid_tids) > 0:
            n_vals = np.array([counts[tid] for tid in valid_tids])[:, None]
            sums_arr = np.array([sums[tid] for tid in valid_tids])
            sums_sq_arr = np.array([sums_sq[tid] for tid in valid_tids])

            mean_matrix = sums_arr / n_vals
            var_matrix = (sums_sq_arr / n_vals) - (mean_matrix ** 2)
            var_matrix = np.maximum(var_matrix, 0)
            std_matrix = np.sqrt(var_matrix)
        else:
            mean_matrix = np.empty((0, nbands))
            std_matrix = np.empty((0, nbands))

        feature_data = {'crop_id': crop_ids, 'seg_id': seg_ids}
        for b in range(nbands):
            feature_data[f'meanB{b}'] = mean_matrix[:, b]
            feature_data[f'stdB{b}'] = std_matrix[:, b]

        df_final = pd.DataFrame(feature_data)
        df_final.to_csv(self.sel_csv, index=False)
        print(f"    Object-Based Features saved to {self.sel_csv}\n")

    # --- Stage 4: Train Classifier ---
    def stage_4_train_classifier(self, **kwargs):
        self._ensure_directories()
        params = self.stage4_params.copy()
        params.update(kwargs)
        stage = 4

        if not self.sel_csv.exists():
            print("ERROR: Feature CSV not found.")
            return

        model_fn = self.model_dir / f"{self.country}_{self.track}_model.pkl"

        print(f"[Stage {stage}/{self.total_stages}] Training ANN...")

        df = pd.read_csv(self.sel_csv)
        feat_cols = [c for c in df.columns if c.startswith('meanB') or c.startswith('stdB')]
        self.feat_cols = feat_cols

        print("    Balancing classes (Capped Oversampling)...")
        threshold = params.get('balance_threshold', 1000)
        df_balanced = pd.DataFrame()
        for crop_id in df['crop_id'].unique():
            df_class = df[df['crop_id'] == crop_id]
            count = len(df_class)

            if count < threshold:
                df_resampled = resample(df_class, replace=True, n_samples=threshold, random_state=42)
                df_balanced = pd.concat([df_balanced, df_resampled])
            else:
                df_balanced = pd.concat([df_balanced, df_class])

        print(f"    Original samples: {len(df)}. Balanced samples: {len(df_balanced)}")

        X = df_balanced[feat_cols].values
        y = df_balanced['crop_id'].values

        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X)

        hidden_sizes = tuple(map(int, str(params['sk_hidden_sizes']).split(',')))

        clf = MLPClassifier(
            hidden_layer_sizes=hidden_sizes,
            activation=params['sk_activation'],
            solver=params['sk_solver'],
            alpha=params['sk_alpha'],
            max_iter=params['sk_max_iter'],
            random_state=42,
            verbose=True
        )

        clf.fit(X_scaled, y)

        joblib.dump({'model': clf, 'scaler': scaler, 'feats': feat_cols}, model_fn)
        print(f"Model saved to {model_fn}")

        y_pred = clf.predict(X_scaled)
        labels = sorted(list(set(y)))
        cm = confusion_matrix(y, y_pred, labels=labels)
        print("\n--- Training Confusion Matrix ---")
        print(pd.DataFrame(cm, index=labels, columns=labels).to_string())
        print("\n")

    # --- Stage 5: Tiled Inference (Object-Based) ---
    def stage_5_classify_vector(self, force_recompute=False):
        # Renamed logic, kept name for compatibility
        self._ensure_directories()
        stage = 5

        model_file = self.model_dir / f"{self.country}_{self.track}_model.pkl"
        if not model_file.exists():
            print("ERROR: Model not found.")
            return

        if self.class_tif.exists() and not force_recompute:
            print(f"[Stage {stage}] Classification Raster exists, skipping.")
            return

        print(f"[Stage {stage}/{self.total_stages}] Running Tiled Object-Based Inference (Parallelized)...")

        data = joblib.load(model_file)
        clf = data['model']
        scaler = data['scaler']
        feat_cols = data['feats']

        ds_stack_info = gdal.Open(str(self.ras))
        cols = ds_stack_info.RasterXSize
        rows = ds_stack_info.RasterYSize
        nbands = ds_stack_info.RasterCount
        gt = ds_stack_info.GetGeoTransform()
        proj = ds_stack_info.GetProjection()
        ds_stack_info = None

        driver = gdal.GetDriverByName('GTiff')
        ds_cls = driver.Create(str(self.class_tif), cols, rows, 1, gdal.GDT_Int32,
                               options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
        ds_cls.SetGeoTransform(gt)
        ds_cls.SetProjection(proj)
        ds_cls.GetRasterBand(1).SetNoDataValue(0)

        ds_conf = driver.Create(str(self.conf_tif), cols, rows, 1, gdal.GDT_Float32,
                                options=['COMPRESS=DEFLATE', 'TILED=YES', 'BIGTIFF=YES'])
        ds_conf.SetGeoTransform(gt)
        ds_conf.SetProjection(proj)
        ds_conf.GetRasterBand(1).SetNoDataValue(0)

        tile_size = 2048

        # We need a lock when writing to the output datasets
        write_lock = threading.Lock()

        def process_tile(x, y):
            xsize = min(tile_size, cols - x)
            ysize = min(tile_size, rows - y)

            # Each thread needs its own GDAL dataset handles to be thread-safe
            ds_stack = gdal.Open(str(self.ras))
            ds_seg = gdal.Open(str(self.seg_tif))

            try:
                seg_arr = ds_seg.GetRasterBand(1).ReadAsArray(x, y, xsize, ysize)
                if np.all(seg_arr == 0):
                    return

                img_list = []
                for b in range(1, nbands + 1):
                    band = ds_stack.GetRasterBand(b)
                    arr = band.ReadAsArray(x, y, xsize, ysize)
                    arr = np.nan_to_num(arr)
                    img_list.append(arr)
                img = np.dstack(img_list)

                flat_seg = seg_arr.ravel()
                mask = flat_seg > 0
                valid_seg = flat_seg[mask]

                if len(valid_seg) == 0:
                    return

                flat_img = img.reshape(-1, nbands)[mask]

                df_img = pd.DataFrame(flat_img, columns=[f'B{i}' for i in range(nbands)])
                df_img['label'] = valid_seg

                grouped = df_img.groupby('label')
                means = grouped.mean()
                stds = grouped.std().fillna(0)

                df_props = pd.DataFrame({'label': means.index})
                for i in range(nbands):
                    df_props[f'meanB{i}'] = means[f'B{i}'].values
                    df_props[f'stdB{i}'] = stds[f'B{i}'].values

                X_tile = df_props[feat_cols].values
                X_scaled = scaler.transform(X_tile)
                preds = clf.predict(X_scaled)
                probs = np.max(clf.predict_proba(X_scaled), axis=1)

                unique_ids = df_props['label'].values
                sort_idx = np.argsort(unique_ids)
                sorted_ids = unique_ids[sort_idx]
                sorted_preds = preds[sort_idx]
                sorted_probs = probs[sort_idx]

                idx_map = np.searchsorted(sorted_ids, valid_seg)
                idx_map = np.clip(idx_map, 0, len(sorted_ids) - 1)
                valid_match = sorted_ids[idx_map] == valid_seg

                flat_cls = np.zeros_like(flat_seg, dtype=np.int32)
                flat_conf = np.zeros_like(flat_seg, dtype=np.float32)

                global_mask = np.zeros(len(flat_seg), dtype=bool)
                global_mask[mask] = valid_match

                flat_cls[global_mask] = sorted_preds[idx_map[valid_match]]
                flat_conf[global_mask] = sorted_probs[idx_map[valid_match]]

                cls_tile = flat_cls.reshape(ysize, xsize)
                conf_tile = flat_conf.reshape(ysize, xsize)

                with write_lock:
                    print(f"    Writing Tile: x={x}, y={y}")
                    ds_cls.GetRasterBand(1).WriteArray(cls_tile, x, y)
                    ds_conf.GetRasterBand(1).WriteArray(conf_tile, x, y)
            finally:
                ds_stack = None
                ds_seg = None


        tiles_to_process = []
        for y in range(0, rows, tile_size):
            for x in range(0, cols, tile_size):
                tiles_to_process.append((x, y))

        print(f"    Dispatching {len(tiles_to_process)} tiles to 20 workers...")
        with ThreadPoolExecutor(max_workers=20) as executor:
            futures = [executor.submit(process_tile, x, y) for x, y in tiles_to_process]
            for future in futures:
                future.result() # Wait for completion and raise any exceptions

        ds_cls.FlushCache()
        ds_conf.FlushCache()
        print(f"    Classification saved to {self.class_tif}\n")

    # --- Stage 6: Mask Class ---
    def stage_6_mask_class(self, force_recompute=False):
        self._ensure_directories()
        stage = 6
        mask_file = self.aux_dir / 'raster_files' / 'EU_arable_areas_mask_3857.tif'
        if not self.class_tif.exists():
            print(f"ERROR: Classified TIF not found.")
            return

        if not self.masked_class.exists() or force_recompute:
            self._apply_mask(self.class_tif, mask_file, self.masked_class, stage)
        else:
            print(f"[Stage {stage}/{self.total_stages}] Masked classification exists, skipping\n")

    # --- Stage 7: Mask Confidence ---
    def stage_7_mask_confidence(self, force_recompute=False):
        self._ensure_directories()
        stage = 7
        mask_file = self.aux_dir / 'raster_files' / 'EU_arable_areas_mask_3857.tif'
        if not self.conf_tif.exists():
            print(f"ERROR: Confidence TIF not found.")
            return

        if not self.masked_conf.exists() or force_recompute:
            self._apply_mask(self.conf_tif, mask_file, self.masked_conf, stage)
        else:
            print(f"[Stage {stage}/{self.total_stages}] Masked confidence exists, skipping\n")

    # --- Stage 8: Metrics ---
    def stage_8_calculate_metrics(self):
        self._ensure_directories()
        stage = 8
        if not self.metrics_fp.exists():
            print(f"[Stage {stage}/{self.total_stages}] Computing metrics...")

            if not self.control_shp.exists():
                print(f"ERROR: Control shapefile not found.")
                return
            if not self.masked_class.exists():
                print(f"ERROR: Masked classification not found.")
                return

            ctrl = gpd.read_file(str(self.control_shp), engine="pyogrio")

            ds = gdal.Open(str(self.masked_class))
            raster_proj = ds.GetProjection()
            raster_srs = osr.SpatialReference()
            raster_srs.ImportFromWkt(raster_proj)

            if ctrl.crs:
                if ctrl.crs.to_wkt() != raster_srs.ExportToWkt():
                    print("    Aligning control points CRS to match raster...")
                    try:
                        from pyproj import CRS
                        target_crs = CRS.from_wkt(raster_srs.ExportToWkt())
                        ctrl = ctrl.to_crs(target_crs)
                    except Exception as e:
                        print(f"    Could not auto-align CRS: {e}")

            band = ds.GetRasterBand(1)
            gt = ds.GetGeoTransform()
            inv = gdal.InvGeoTransform(gt)
            true_vals, pred_vals = [], []

            xs = ctrl.geometry.x.values
            ys = ctrl.geometry.y.values

            pxs = (inv[0] + inv[1] * xs + inv[2] * ys).astype(int)
            pys = (inv[3] + inv[4] * xs + inv[5] * ys).astype(int)
            crop_ids = ctrl['crop_id'].values

            for px, py, crop_id in zip(pxs, pys, crop_ids):
                try:
                    if 0 <= px < ds.RasterXSize and 0 <= py < ds.RasterYSize:
                        t = int(crop_id)
                        val_arr = band.ReadAsArray(px, py, 1, 1)
                        if val_arr is not None:
                            p = int(val_arr[0, 0])
                            if t > 0 and p > 0 and p != -9999:
                                true_vals.append(t)
                                pred_vals.append(p)
                except Exception as e:
                    print(f"    [WARNING] Failed to extract point value: {e}")

            if not true_vals or not pred_vals:
                print("ERROR: No valid matching true/predicted values found.")
                print("HINT: Ensure your test points intersect valid data areas in the masked raster.")
                return

            labels = sorted(list(set(true_vals + pred_vals)))
            cm = confusion_matrix(true_vals, pred_vals, labels=labels)
            precisions, recalls, f1s, _ = precision_recall_fscore_support(
                true_vals, pred_vals, labels=labels, average=None, zero_division=0
            )

            total = np.sum(cm)
            oa = np.trace(cm) / total
            sum_po = oa
            sum_pe = np.sum(np.sum(cm, axis=0) * np.sum(cm, axis=1)) / (total ** 2)
            kappa = (sum_po - sum_pe) / (1 - sum_pe) if (1 - sum_pe) != 0 else np.nan

            resx, resy = abs(gt[1]), abs(gt[5])
            area_ha = resx * resy / 10000

            arr = band.ReadAsArray()
            if arr is not None:
                unique_classes, counts = np.unique(arr[arr > 0], return_counts=True)
                class_areas = dict(zip(unique_classes, counts))
                areas = [{'Class': c, 'Area_ha': round(class_areas.get(c, 0) * area_ha, 2)} for c in labels]
            else:
                areas = [{'Class': c, 'Area_ha': 0} for c in labels]

            wb = openpyxl.Workbook()
            sh = wb.active
            sh.title = 'Results'

            sh.cell(row=1, column=1, value='Confusion Matrix').font = Font(bold=True)
            sh.cell(row=2, column=1, value='True \\ Pred').font = Font(bold=True)
            for j, lbl in enumerate(labels, start=2):
                sh.cell(row=2, column=j, value=lbl).font = Font(bold=True)
            for i, lbl in enumerate(labels, start=3):
                sh.cell(row=i, column=1, value=lbl).font = Font(bold=True)
                for j, _ in enumerate(labels):
                    sh.cell(row=i, column=j + 2, value=int(cm[i - 3, j]))

            base = 4 + len(labels)
            sh.cell(row=base, column=1, value='Overall Accuracy').font = Font(bold=True)
            sh.cell(row=base, column=2, value=round(oa, 4))
            sh.cell(row=base + 1, column=1, value='Kappa').font = Font(bold=True)
            sh.cell(row=base + 1, column=2, value=round(kappa, 4))

            start = base + 3
            headers = ['Class', 'Producer Acc (Recall)', 'User Acc (Precision)', 'F1-score']
            for j, h in enumerate(headers, start=1):
                sh.cell(row=start, column=j, value=h).font = Font(bold=True)
            for idx, c in enumerate(labels):
                row_idx = start + 1 + idx
                sh.cell(row=row_idx, column=1, value=c)
                sh.cell(row=row_idx, column=2, value=round(recalls[idx], 4))
                sh.cell(row=row_idx, column=3, value=round(precisions[idx], 4))
                sh.cell(row=row_idx, column=4, value=round(f1s[idx], 4))

            ar0 = start + 1 + len(labels) + 1
            sh.cell(row=ar0, column=1, value='Areas (ha)').font = Font(bold=True)
            sh.cell(row=ar0 + 1, column=1, value='Class').font = Font(bold=True)
            sh.cell(row=ar0 + 1, column=2, value='Area_ha').font = Font(bold=True)
            for idx, a in enumerate(areas, start=ar0 + 2):
                sh.cell(row=idx, column=1, value=a['Class'])
                sh.cell(row=idx, column=2, value=a['Area_ha'])

            wb.save(str(self.metrics_fp))
            print(f"Metrics saved to {self.metrics_fp}\n")
        else:
            print(f"[Stage 8] Metrics Excel exists, skipping")


# --- Interactive Menu Helpers ---

def get_params(param_dict):
    new_params = param_dict.copy()
    print("--- Current Parameters ---")
    for key, val in new_params.items():
        print(f"  {key}: {val}")

    if input("Change parameters? (y/n) [n]: ").lower() != 'y':
        return new_params

    for key, val in new_params.items():
        new_val_str = input(f"Enter new value for '{key}' [{val}]: ")
        if not new_val_str:
            continue
        try:
            original_type = type(val)
            new_params[key] = original_type(new_val_str)
        except ValueError:
            print(f"Invalid value. Keeping default {val}.")
    return new_params


def get_classifier_params(param_dict):
    new_params = param_dict.copy()
    print("--- Current Parameters ---")
    for key, val in new_params.items():
        print(f"  {key}: {val}")

    if input("Change parameters? (y/n) [n]: ").lower() != 'y':
        return new_params

    clf = input(f"Enter classifier (ann_sklearn) [{new_params['classifier']}]: ") or new_params['classifier']
    new_params['classifier'] = clf.lower()

    print(f"\n--- Setting parameters for {clf.upper()} ---")
    prefix = 'sk_'
    for key in [k for k in new_params if k.startswith(prefix)]:
        val = new_params[key]
        new_val_str = input(f"Enter new value for '{key}' [{val}]: ")
        if new_val_str:
            try:
                new_params[key] = type(val)(new_val_str)
            except ValueError:
                print(f"Invalid value.")
    return new_params


# --- Main Execution ---

def main_menu(pipeline):
    menu = f"""
    --- Raster-Based OBIA Pipeline (ANN) ---
    Track: {pipeline.track} ({pipeline.country})

    [1] Stage 1: OTB Segmentation (Raster Mode)
    [2] Stage 2: Split Samples
    [3] Stage 3: Extract Features (Object-based Training)
    [4] Stage 4: Train ANN Classifier
    [5] Stage 5: Tiled Object-Based Inference
    [6] Stage 6: Mask Classification
    [7] Stage 7: Mask Confidence
    [8] Stage 8: Calculate Metrics

    [A] Run All Stages (Forces overwrite of Stages 5-8 to clear old bugs)
    [Q] Quit

    Enter your choice:
    """

    while True:
        choice = input(menu).strip().upper()
        try:
            if choice == '1':
                new_params = get_params(pipeline.stage1_params)
                pipeline.stage1_params.update(new_params)
                pipeline.stage_1_segmentation(**pipeline.stage1_params)
            elif choice == '2':
                new_params = get_params(pipeline.stage2_params)
                pipeline.stage2_params.update(new_params)
                pipeline.stage_2_split_samples(**pipeline.stage2_params)
            elif choice == '3':
                pipeline.stage_3_selection()
            elif choice == '4':
                new_params = get_classifier_params(pipeline.stage4_params)
                force = (pipeline.stage4_params != new_params)
                pipeline.stage4_params.update(new_params)
                pipeline.stage_4_train_classifier(force_retrain=force, **pipeline.stage4_params)
            elif choice == '5':
                pipeline.stage_5_classify_vector()
            elif choice == '6':
                pipeline.stage_6_mask_class(force_recompute=True)
            elif choice == '7':
                pipeline.stage_7_mask_confidence(force_recompute=True)
            elif choice == '8':
                pipeline.stage_8_calculate_metrics()
            elif choice == 'A':
                print(
                    "\nNOTE: Running 'A' will automatically force recomputation of Stages 5-8 to clear any corrupted old files.")
                pipeline.stage_1_segmentation(**pipeline.stage1_params)
                pipeline.stage_2_split_samples(**pipeline.stage2_params)
                pipeline.stage_3_selection()
                pipeline.stage_4_train_classifier(**pipeline.stage4_params)
                # Force recompute inference and masking
                pipeline.stage_5_classify_vector(force_recompute=True)
                pipeline.stage_6_mask_class(force_recompute=True)
                pipeline.stage_7_mask_confidence(force_recompute=True)
                if pipeline.metrics_fp.exists(): pipeline.metrics_fp.unlink()
                pipeline.stage_8_calculate_metrics()
            elif choice == 'Q':
                break
        except Exception as e:
            print(f"\n--- ERROR ---: {e}")
            import traceback
            traceback.print_exc()


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Modular OBIA Pipeline (True Object-Based Training)")
    parser.add_argument('--track', required=True, help="Processing track ID (e.g., P1, P2)")
    args = parser.parse_args()

    try:
        pipeline = ProcessingPipeline(track=args.track)
        main_menu(pipeline)
    except Exception as e:
        print(f"Initialization Error: {e}")
        sys.exit(1)